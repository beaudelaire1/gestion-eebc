import io
from datetime import date

import pytest
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.accounts.models import User
from apps.core.models import AuditLog
from apps.imports.models import ImportLog
from apps.imports.services import ExcelImportService, generate_template_excel
from apps.members.models import Member
from apps.young.models import YoungMember, YouthGroup
from test_factories import MemberFactory, UserFactory

pytestmark = pytest.mark.django_db


@pytest.fixture
def run_import(admin_user):
    def run(**overrides):
        values = {
            "prenom": "Camille",
            "nom": "Exemple",
            "date_naissance": "15/06/2007",
            "email": "camille@example.com",
            "telephone": "0694123456",
        }
        values.update(overrides)
        workbook = Workbook()
        workbook.active.append(list(values))
        workbook.active.append(list(values.values()))
        data = io.BytesIO()
        workbook.save(data)
        log = ImportLog.objects.create(
            import_type=ImportLog.ImportType.YOUNG_MEMBERS,
            imported_by=admin_user,
            file_name="young.xlsx",
            file_path=SimpleUploadedFile("young.xlsx", data.getvalue()),
        )
        service = ExcelImportService(log)
        service.process_import()
        log.refresh_from_db()
        return log

    return run


def test_legacy_import_preserves_phone_and_does_not_create_links(
    run_import, admin_user
):
    log = run_import()
    young = YoungMember.objects.get()
    assert log.success_rows == 1
    assert young.phone == "0694123456"
    assert young.linked_member_id is None and young.user_id is None
    assert Member.objects.count() == 0
    assert User.objects.count() == 1


def test_church_member_and_account_are_created_once(run_import, admin_user, mailoutbox):
    for _ in range(2):
        log = run_import(membre_eglise="oui", creer_compte="oui")
        assert log.error_rows == 0
    young = YoungMember.objects.get()
    assert Member.objects.count() == 1 and User.objects.count() == 2
    assert young.linked_member.user_id == young.user_id
    assert young.user.role == "membre"
    assert young.user.created_by_id == admin_user.pk
    assert young.user.must_change_password
    assert not young.user.is_staff and not young.user.is_superuser
    assert len(mailoutbox) == 0
    assert AuditLog.objects.filter(
        model_name="young.YoungMember", user=admin_user
    ).exists()


def test_account_can_be_created_without_church_membership(run_import):
    log = run_import(membre_eglise="non", creer_compte="oui")
    assert log.error_rows == 0
    young = YoungMember.objects.get()
    assert young.user_id and not young.linked_member_id
    assert not Member.objects.exists()


def test_existing_member_and_account_are_reused_without_changing_password_or_role(
    run_import,
):
    user = UserFactory(role="encadrant")
    member = MemberFactory(
        first_name="Camille",
        last_name="Exemple",
        date_of_birth=date(2007, 6, 15),
        user=user,
    )
    password = user.password
    log = run_import(membre_eglise="oui", creer_compte="oui")
    assert log.error_rows == 0
    young = YoungMember.objects.get()
    assert young.linked_member_id == member.pk and young.user_id == user.pk
    user.refresh_from_db()
    assert user.password == password and user.role == "encadrant"


def test_existing_account_is_not_claimed_by_email(run_import):
    UserFactory(email="camille@example.com")
    log = run_import(membre_eglise="oui", creer_compte="oui")
    assert log.error_rows == 1 and log.success_rows == 0
    assert not YoungMember.objects.exists() and not Member.objects.exists()
    assert "email" in log.error_log


@pytest.mark.parametrize("choice", ["peut-être", "1", "true"])
def test_invalid_decision_rolls_back_entire_row(run_import, choice):
    log = run_import(membre_eglise="oui", creer_compte=choice)
    assert log.error_rows == 1
    assert not YoungMember.objects.exists() and not Member.objects.exists()


@pytest.mark.parametrize("email", ["", "incorrect"])
def test_account_requires_valid_email(run_import, email):
    log = run_import(email=email, creer_compte="oui")
    assert log.error_rows == 1 and not YoungMember.objects.exists()


def test_ambiguous_member_requires_explicit_id(run_import):
    candidates = [
        MemberFactory(
            first_name="Camille", last_name="Exemple", date_of_birth=date(2007, 6, 15)
        )
        for _ in range(2)
    ]
    assert run_import(membre_eglise="oui").error_rows == 1
    assert not YoungMember.objects.exists()
    assert (
        run_import(membre_eglise="oui", membre_id=candidates[1].member_id).error_rows
        == 0
    )
    assert YoungMember.objects.get().linked_member_id == candidates[1].pk


def test_member_without_birth_date_requires_id(run_import):
    member = MemberFactory(
        first_name="Camille", last_name="Exemple", date_of_birth=None
    )
    assert run_import(membre_eglise="oui").error_rows == 1
    assert run_import(membre_eglise="oui", membre_id=member.member_id).error_rows == 0


def test_wrong_member_id_is_rejected(run_import):
    member = MemberFactory(first_name="Autre", last_name="Personne")
    assert run_import(membre_eglise="oui", membre_id=member.member_id).error_rows == 1
    assert not YoungMember.objects.exists()


def test_no_does_not_unlink_existing_member(run_import):
    assert run_import(membre_eglise="oui").error_rows == 0
    original = YoungMember.objects.get().linked_member_id
    assert run_import(membre_eglise="non").error_rows == 1
    assert YoungMember.objects.get().linked_member_id == original


def test_existing_young_is_rolled_back_on_link_failure(run_import):
    assert run_import(telephone="").error_rows == 0
    assert (
        run_import(
            telephone="0694000000", membre_id="UNKNOWN", membre_eglise="oui"
        ).error_rows
        == 1
    )
    assert YoungMember.objects.get().phone == ""


def test_age_group_does_not_advance_before_birthday(run_import):
    today = date.today()
    # Keep the birthday tomorrow without depending on leap-year replacement.
    from datetime import timedelta

    tomorrow = today + timedelta(days=1)
    dob = (
        date(tomorrow.year - 18, tomorrow.month, tomorrow.day)
        if tomorrow.strftime("%m%d") != "0229"
        else None
    )
    if tomorrow.day == 29 and tomorrow.month == 2:
        dob = date(tomorrow.year - 18, 3, 1)
    YouthGroup.objects.create(name="Mineurs", min_age=13, max_age=17)
    YouthGroup.objects.create(name="Adultes", min_age=18, max_age=30)
    assert run_import(date_naissance=dob.strftime("%d/%m/%Y")).error_rows == 0
    assert YoungMember.objects.get().group.name == "Mineurs"


def test_template_download_contains_matching_instructions(client, admin_user):
    client.force_login(admin_user)
    response = client.get(reverse("imports:template", args=["young_members"]))
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    headers = [c.value for c in workbook["Données"][1]]
    assert headers == list(generate_template_excel("young_members").columns)
    assert headers[-3:] == ["membre_eglise", "membre_id", "creer_compte"]
    assert workbook["Instructions"].max_row == len(headers) + 1


def test_member_cannot_import(client, member_user):
    client.force_login(member_user)
    response = client.post(reverse("imports:create"))
    assert response.status_code in (302, 403)
    assert not ImportLog.objects.exists()


def test_model_rejects_different_member_and_young_accounts():
    member = MemberFactory(user=UserFactory())
    young = YoungMember(linked_member=member, user=UserFactory())
    with pytest.raises(ValidationError):
        young.clean()
