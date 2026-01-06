"""
Services génériques pour l'export et l'impression.

Ce module fournit des services réutilisables pour :
- Export Excel avec formatage automatique
- Génération de PDF pour impression
- Templates d'impression standardisés
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Service d'export générique pour Excel et PDF."""
    
    # Styles Excel prédéfinis
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="0A36FF", end_color="0A36FF", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=16, color="0B0F19")
    SUBTITLE_FONT = Font(bold=True, size=12, color="475569")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @classmethod
    def export_to_excel(
        cls,
        data: List[Dict[str, Any]],
        headers: List[str],
        title: str,
        subtitle: str = "",
        filename: str = None,
        metadata: Dict[str, Any] = None
    ) -> HttpResponse:
        """
        Exporte des données vers Excel avec formatage automatique.
        
        Args:
            data: Liste de dictionnaires contenant les données
            headers: Liste des en-têtes de colonnes
            title: Titre principal du document
            subtitle: Sous-titre optionnel
            filename: Nom du fichier (généré automatiquement si None)
            metadata: Métadonnées additionnelles à afficher
        
        Returns:
            HttpResponse avec le fichier Excel
        """
        try:
            # Créer le classeur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"
            
            current_row = 1
            
            # Titre principal
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = title
            title_cell.font = cls.TITLE_FONT
            title_cell.alignment = Alignment(horizontal='center')
            current_row += 2
            
            # Sous-titre
            if subtitle:
                ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
                subtitle_cell = ws[f'A{current_row}']
                subtitle_cell.value = subtitle
                subtitle_cell.font = cls.SUBTITLE_FONT
                subtitle_cell.alignment = Alignment(horizontal='center')
                current_row += 2
            
            # Métadonnées
            if metadata:
                for key, value in metadata.items():
                    ws[f'A{current_row}'] = f"{key}:"
                    ws[f'A{current_row}'].font = Font(bold=True)
                    ws[f'B{current_row}'] = str(value)
                    current_row += 1
                current_row += 1
            
            # Date d'export
            ws[f'A{current_row}'] = "Date d'export:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            current_row += 2
            
            # En-têtes
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = cls.HEADER_FONT
                cell.fill = cls.HEADER_FILL
                cell.border = cls.BORDER
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Données
            for row_data in data:
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    # Récupérer la valeur par clé ou par index
                    if isinstance(row_data, dict):
                        cell.value = row_data.get(header, "")
                    else:
                        cell.value = row_data[col_num - 1] if col_num - 1 < len(row_data) else ""
                    cell.border = cls.BORDER
                    
                    # Formatage automatique des nombres
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                        if '€' in str(cell.value) or 'montant' in header.lower():
                            cell.number_format = '#,##0.00 "€"'
                
                current_row += 1
            
            # Ajuster la largeur des colonnes
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15
            
            # Générer le nom de fichier
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_{timestamp}.xlsx"
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export Excel: {e}")
            raise
    
    @classmethod
    def prepare_print_context(
        cls,
        title: str,
        data: Any,
        template_name: str,
        extra_context: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """
        Prépare le contexte pour l'impression.
        
        Args:
            title: Titre du document
            data: Données principales
            template_name: Nom du template d'impression
            extra_context: Contexte additionnel
        
        Returns:
            Contexte complet pour le template
        """
        context = {
            'title': title,
            'data': data,
            'print_date': datetime.now(),
            'print_timestamp': datetime.now().strftime('%d/%m/%Y à %H:%M'),
            'site_name': getattr(settings, 'SITE_NAME', 'EEBC'),
            'is_print_view': True,
        }
        
        if extra_context:
            context.update(extra_context)
        
        return context


class PrintService:
    """Service d'impression générique."""
    
    @classmethod
    def render_print_template(
        cls,
        request,
        template_name: str,
        context: Dict[str, Any]
    ) -> HttpResponse:
        """
        Rend un template optimisé pour l'impression.
        
        Args:
            request: Requête HTTP
            template_name: Nom du template
            context: Contexte du template
        
        Returns:
            HttpResponse avec le HTML d'impression
        """
        # Ajouter les styles d'impression au contexte
        context['print_styles'] = cls._get_print_styles()
        
        return render_to_string(template_name, context, request=request)
    
    @classmethod
    def _get_print_styles(cls) -> str:
        """Retourne les styles CSS optimisés pour l'impression."""
        return """
        <style>
            @media print {
                body { 
                    font-family: 'Arial', sans-serif;
                    font-size: 12pt;
                    line-height: 1.4;
                    color: #000;
                    background: white;
                }
                
                .no-print { display: none !important; }
                
                .print-header {
                    border-bottom: 2px solid #0A36FF;
                    padding-bottom: 1rem;
                    margin-bottom: 2rem;
                }
                
                .print-title {
                    font-size: 18pt;
                    font-weight: bold;
                    color: #0A36FF;
                    margin: 0;
                }
                
                .print-subtitle {
                    font-size: 12pt;
                    color: #666;
                    margin: 0.5rem 0 0 0;
                }
                
                .print-meta {
                    font-size: 10pt;
                    color: #888;
                    margin-top: 1rem;
                }
                
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 1rem 0;
                }
                
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                
                th {
                    background-color: #f5f5f5;
                    font-weight: bold;
                }
                
                .text-right { text-align: right; }
                .text-center { text-align: center; }
                
                .page-break {
                    page-break-before: always;
                }
                
                .print-footer {
                    position: fixed;
                    bottom: 0;
                    width: 100%;
                    text-align: center;
                    font-size: 10pt;
                    color: #666;
                    border-top: 1px solid #ddd;
                    padding-top: 0.5rem;
                }
            }
            
            @media screen {
                .print-preview {
                    max-width: 21cm;
                    margin: 2rem auto;
                    padding: 2rem;
                    background: white;
                    box-shadow: 0 0 20px rgba(0,0,0,0.1);
                }
                
                .print-actions {
                    text-align: center;
                    margin: 2rem 0;
                }
                
                .btn-print {
                    background: #0A36FF;
                    color: white;
                    border: none;
                    padding: 0.75rem 2rem;
                    border-radius: 8px;
                    font-weight: 600;
                    cursor: pointer;
                    margin: 0 0.5rem;
                }
                
                .btn-print:hover {
                    background: #0829CC;
                }
            }
        </style>
        """


class DataExportMixin:
    """
    Mixin pour ajouter facilement les fonctionnalités d'export aux vues.
    """
    
    export_filename_prefix = "export"
    export_title = "Export de données"
    export_headers = []
    
    def get_export_data(self):
        """
        Méthode à surcharger pour retourner les données à exporter.
        Doit retourner une liste de dictionnaires.
        """
        raise NotImplementedError("Vous devez implémenter get_export_data()")
    
    def get_export_headers(self):
        """Retourne les en-têtes pour l'export."""
        return self.export_headers
    
    def get_export_filename(self):
        """Génère le nom de fichier pour l'export."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{self.export_filename_prefix}_{timestamp}.xlsx"
    
    def get_export_metadata(self):
        """Retourne les métadonnées pour l'export."""
        return {
            'Utilisateur': self.request.user.get_full_name() or self.request.user.username,
            'Application': self.__class__.__module__.split('.')[1].title(),
        }
    
    def export_excel(self, request):
        """Exporte les données en Excel."""
        data = self.get_export_data()
        headers = self.get_export_headers()
        
        return ExportService.export_to_excel(
            data=data,
            headers=headers,
            title=self.export_title,
            filename=self.get_export_filename(),
            metadata=self.get_export_metadata()
        )


class PrintMixin:
    """
    Mixin pour ajouter facilement les fonctionnalités d'impression aux vues.
    """
    
    print_template_name = None
    print_title = "Document d'impression"
    
    def get_print_template_name(self):
        """Retourne le nom du template d'impression."""
        if self.print_template_name:
            return self.print_template_name
        
        # Génération automatique basée sur le nom de la vue
        app_name = self.__class__.__module__.split('.')[1]
        view_name = self.__class__.__name__.lower().replace('view', '')
        return f"{app_name}/{view_name}_print.html"
    
    def get_print_context(self):
        """Retourne le contexte pour l'impression."""
        return {
            'object': getattr(self, 'object', None),
            'object_list': getattr(self, 'object_list', None),
        }
    
    def print_view(self, request, *args, **kwargs):
        """Vue d'impression."""
        context = ExportService.prepare_print_context(
            title=self.print_title,
            data=self.get_print_context(),
            template_name=self.get_print_template_name(),
            extra_context={'request': request}
        )
        
        return render_to_string(
            self.get_print_template_name(),
            context,
            request=request
        )