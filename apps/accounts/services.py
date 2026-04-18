"""
Service d'authentification centralisé pour Gestion EEBC.
Gère le flux de login sécurisé, les tokens de changement de mot de passe,
le rate limiting et la création d'utilisateurs.
"""
import secrets
import string
import unicodedata
from django.contrib.auth import authenticate
from django.core import signing
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from django.utils.html import strip_tags
from datetime import timedelta
from typing import Optional, Tuple

from .models import User, PasswordChangeToken


class ServiceResult:
    """Résultat d'une opération de service."""
    
    def __init__(self, success: bool, data=None, error: str = None):
        self.success = success
        self.data = data
        self.error = error
    
    @classmethod
    def ok(cls, data=None):
        return cls(success=True, data=data)
    
    @classmethod
    def fail(cls, error: str):
        return cls(success=False, error=error)


class AuthenticationService:
    """Service d'authentification centralisé."""
    
    # Configuration par défaut
    TOKEN_EXPIRY_MINUTES = 30
    MAX_LOGIN_ATTEMPTS = 5
    LOCKOUT_MINUTES = 15
    SIGNING_SALT = 'password_change_token'
    
    @classmethod
    def authenticate_user(
        cls,
        username: str,
        password: str,
        request=None
    ) -> Tuple[Optional[User], str]:
        """
        Authentifie un utilisateur avec rate limiting et logging.
        
        Args:
            username: Nom d'utilisateur
            password: Mot de passe
            request: Requête HTTP (optionnel, pour l'IP)
        
        Returns:
            Tuple (user, error_message)
            - Si succès: (user, "")
            - Si échec: (None, message d'erreur)
        """
        # Vérifier si l'utilisateur existe
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return None, "Nom d'utilisateur ou mot de passe incorrect."
        
        # Vérifier si le compte est verrouillé
        if user.is_locked():
            remaining = user.locked_until - timezone.now()
            minutes = int(remaining.total_seconds() / 60) + 1
            return None, f"Compte temporairement verrouillé. Réessayez dans {minutes} minute(s)."
        
        # Tenter l'authentification
        authenticated_user = authenticate(username=username, password=password)
        
        if authenticated_user is None:
            # Enregistrer l'échec
            user.record_failed_attempt(
                lockout_minutes=cls.LOCKOUT_MINUTES,
                max_attempts=cls.MAX_LOGIN_ATTEMPTS
            )
            
            remaining_attempts = cls.MAX_LOGIN_ATTEMPTS - user.failed_login_attempts
            if remaining_attempts > 0:
                return None, f"Nom d'utilisateur ou mot de passe incorrect. {remaining_attempts} tentative(s) restante(s)."
            else:
                return None, f"Compte verrouillé pendant {cls.LOCKOUT_MINUTES} minutes suite à trop de tentatives échouées."
        
        # Succès - réinitialiser les tentatives
        user.reset_failed_attempts()
        
        # Enregistrer l'IP de connexion
        if request:
            ip = cls.get_client_ip(request)
            user.last_login_ip = ip
            user.save(update_fields=['last_login_ip'])
        
        return authenticated_user, ""
    
    @classmethod
    def generate_password_change_token(cls, user: User) -> str:
        """
        Génère un token signé pour le changement de mot de passe.
        
        Args:
            user: L'utilisateur pour lequel générer le token
        
        Returns:
            Le token signé sous forme de chaîne
        """
        # Invalider les anciens tokens non utilisés
        PasswordChangeToken.objects.filter(
            user=user,
            used=False
        ).update(used=True)
        
        # Créer un nouveau token
        raw_token = PasswordChangeToken.generate_token()
        expires_at = timezone.now() + timedelta(minutes=cls.TOKEN_EXPIRY_MINUTES)
        
        token_obj = PasswordChangeToken.objects.create(
            user=user,
            token=raw_token,
            expires_at=expires_at
        )
        
        # Signer le token pour plus de sécurité
        signed_token = signing.dumps(
            {'token': raw_token, 'user_id': user.id},
            salt=cls.SIGNING_SALT
        )
        
        return signed_token
    
    @classmethod
    def verify_password_change_token(cls, signed_token: str) -> Optional[User]:
        """
        Vérifie et retourne l'utilisateur du token.
        
        Args:
            signed_token: Le token signé à vérifier
        
        Returns:
            L'utilisateur si le token est valide, None sinon
        """
        try:
            # Décoder le token signé
            data = signing.loads(
                signed_token,
                salt=cls.SIGNING_SALT,
                max_age=cls.TOKEN_EXPIRY_MINUTES * 60  # en secondes
            )
            
            raw_token = data.get('token')
            user_id = data.get('user_id')
            
            if not raw_token or not user_id:
                return None
            
            # Vérifier le token en base
            token_obj = PasswordChangeToken.objects.filter(
                token=raw_token,
                user_id=user_id,
                used=False
            ).first()
            
            if token_obj and token_obj.is_valid():
                return token_obj.user
            
            return None
            
        except (signing.BadSignature, signing.SignatureExpired):
            return None
    
    @classmethod
    def consume_password_change_token(cls, signed_token: str) -> Optional[User]:
        """
        Vérifie et consomme le token (le marque comme utilisé).
        
        Args:
            signed_token: Le token signé à consommer
        
        Returns:
            L'utilisateur si le token est valide, None sinon
        """
        try:
            data = signing.loads(
                signed_token,
                salt=cls.SIGNING_SALT,
                max_age=cls.TOKEN_EXPIRY_MINUTES * 60
            )
            
            raw_token = data.get('token')
            user_id = data.get('user_id')
            
            if not raw_token or not user_id:
                return None
            
            token_obj = PasswordChangeToken.objects.filter(
                token=raw_token,
                user_id=user_id,
                used=False
            ).first()
            
            if token_obj and token_obj.is_valid():
                token_obj.mark_as_used()
                return token_obj.user
            
            return None
            
        except (signing.BadSignature, signing.SignatureExpired):
            return None
    
    @classmethod
    def record_login_attempt(cls, user: User, success: bool, ip: str = None):
        """
        Enregistre une tentative de connexion.
        
        Args:
            user: L'utilisateur concerné
            success: True si la connexion a réussi
            ip: Adresse IP du client
        """
        if success:
            user.reset_failed_attempts()
            if ip:
                user.last_login_ip = ip
                user.save(update_fields=['last_login_ip'])
        else:
            user.record_failed_attempt(
                lockout_minutes=cls.LOCKOUT_MINUTES,
                max_attempts=cls.MAX_LOGIN_ATTEMPTS
            )
    
    @staticmethod
    def get_client_ip(request) -> Optional[str]:
        """
        Récupère l'adresse IP du client.
        
        Args:
            request: La requête HTTP
        
        Returns:
            L'adresse IP ou None
        """
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    @classmethod
    def is_user_locked(cls, username: str) -> Tuple[bool, Optional[int]]:
        """
        Vérifie si un utilisateur est verrouillé.
        
        Args:
            username: Nom d'utilisateur
        
        Returns:
            Tuple (is_locked, remaining_minutes)
        """
        try:
            user = User.objects.get(username=username)
            if user.is_locked():
                remaining = user.locked_until - timezone.now()
                minutes = int(remaining.total_seconds() / 60) + 1
                return True, minutes
            return False, None
        except User.DoesNotExist:
            return False, None
    
    @classmethod
    def cleanup_expired_tokens(cls):
        """Nettoie les tokens expirés."""
        PasswordChangeToken.objects.filter(
            expires_at__lt=timezone.now()
        ).delete()


class AccountsService:
    """
    Service de gestion des comptes utilisateurs.
    Centralise la logique métier pour la création et l'activation des comptes.
    """
    
    @staticmethod
    def generate_username(first_name: str, last_name: str) -> str:
        """
        Génère un nom d'utilisateur au format: prénom_nom (en minuscules, sans accents).
        Ex: Paul Kapel -> pa_kapel
        
        Args:
            first_name: Prénom de l'utilisateur
            last_name: Nom de l'utilisateur
        
        Returns:
            Un nom d'utilisateur unique
        """
        # Nettoyer et normaliser les noms
        first_clean = unicodedata.normalize('NFD', first_name.lower())
        first_clean = ''.join(c for c in first_clean if unicodedata.category(c) != 'Mn')
        
        last_clean = unicodedata.normalize('NFD', last_name.lower())
        last_clean = ''.join(c for c in last_clean if unicodedata.category(c) != 'Mn')
        
        # Prendre les 2 premières lettres du prénom + nom complet
        first_part = first_clean[:2] if len(first_clean) >= 2 else first_clean
        username = f"{first_part}_{last_clean}"
        
        # Vérifier l'unicité et ajouter un numéro si nécessaire
        original_username = username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{original_username}{counter}"
            counter += 1
        
        return username
    
    @staticmethod
    def generate_password(length: int = 12) -> str:
        """
        Génère un mot de passe sécurisé.
        
        Args:
            length: Longueur du mot de passe (défaut: 12)
        
        Returns:
            Un mot de passe sécurisé
        """
        # Caractères autorisés (éviter les caractères ambigus)
        letters = string.ascii_letters.replace('O', '').replace('l', '').replace('I', '')
        digits = string.digits.replace('0', '')  # Éviter le zéro ambigu
        symbols = "!@#$%&*"
        chars = letters + digits + symbols
        
        password = ''.join(secrets.choice(chars) for _ in range(length))
        
        # S'assurer qu'il y a au moins une majuscule, une minuscule, un chiffre et un symbole
        if not any(c.isupper() for c in password):
            password = password[:-1] + secrets.choice(string.ascii_uppercase.replace('O', '').replace('I', ''))
        if not any(c.islower() for c in password):
            password = password[:-1] + secrets.choice(string.ascii_lowercase.replace('l', ''))
        if not any(c.isdigit() for c in password):
            password = password[:-1] + secrets.choice(digits)
        if not any(c in symbols for c in password):
            password = password[:-1] + secrets.choice(symbols)
        
        return password
    
    @classmethod
    def create_user_by_team(
        cls,
        first_name: str,
        last_name: str,
        email: str,
        roles: list,  # Maintenant une liste de rôles
        created_by: User,
        phone: str = '',
        send_email: bool = True
    ) -> ServiceResult:
        """
        Crée un nouvel utilisateur avec identifiants générés automatiquement.
        
        Args:
            first_name: Prénom
            last_name: Nom
            email: Email
            roles: Liste des rôles (User.Role)
            created_by: Utilisateur qui crée le compte
            phone: Téléphone (optionnel)
            send_email: Envoyer l'email d'invitation (défaut: True)
        
        Returns:
            ServiceResult avec data={'user': user, 'username': username, 'password': password}
            ou ServiceResult.fail() si erreur
        """
        try:
            # Générer les identifiants
            username = cls.generate_username(first_name, last_name)
            password = cls.generate_password()
            
            # Convertir la liste de rôles en chaîne
            role_string = ','.join(roles) if roles else User.Role.MEMBRE
            
            # Créer l'utilisateur
            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=password,
                role=role_string,
                phone=phone,
                created_by_team=True,
                created_by=created_by,
                must_change_password=True  # Forcer le changement de mot de passe
            )
            
            # Envoyer l'email d'invitation si demandé
            if send_email:
                cls.send_invitation_email(user, username, password, created_by)
            
            return ServiceResult.ok({
                'user': user,
                'username': username,
                'password': password
            })
            
        except Exception as e:
            return ServiceResult.fail(f"Erreur lors de la création de l'utilisateur: {str(e)}")
    
    @classmethod
    def activate_user_account(cls, user: User, new_password: str) -> ServiceResult:
        """
        Active le compte utilisateur après le premier changement de mot de passe.
        
        Args:
            user: L'utilisateur à activer
            new_password: Le nouveau mot de passe
        
        Returns:
            ServiceResult.ok() si succès, ServiceResult.fail() si erreur
        """
        try:
            user.set_password(new_password)
            user.must_change_password = False
            user.save(update_fields=['password', 'must_change_password'])
            return ServiceResult.ok({'user': user})
        except Exception as e:
            return ServiceResult.fail(f"Erreur lors de l'activation du compte: {str(e)}")
    
    @classmethod
    def send_invitation_email(
        cls,
        user: User,
        username: str,
        password: str,
        created_by: User
    ) -> bool:
        """
        Envoie l'email d'invitation avec les identifiants.
        
        Args:
            user: L'utilisateur invité
            username: Nom d'utilisateur
            password: Mot de passe temporaire
            created_by: Utilisateur qui a créé le compte
        
        Returns:
            True si l'email a été envoyé, False sinon
        """
        try:
            # Générer un token pour le changement de mot de passe direct
            token = AuthenticationService.generate_password_change_token(user)
            
            context = {
                'user': user,
                'username': username,
                'password': password,
                'created_by': created_by,
                'login_url': f"{getattr(settings, 'SITE_URL', '')}/accounts/login/",
                'password_change_url': f"{getattr(settings, 'SITE_URL', '')}/accounts/first-login-password-change/?token={token}",
                'site_name': getattr(settings, 'SITE_NAME', 'EEBC'),
            }
            
            # Rendu du template email
            html_message = render_to_string('accounts/emails/user_invitation.html', context)
            plain_message = strip_tags(html_message)
            
            # Utiliser l'email Hostinger si configuré, sinon DEFAULT_FROM_EMAIL
            from_email = getattr(settings, 'HOSTINGER_EMAIL_HOST_USER', None) or settings.DEFAULT_FROM_EMAIL
            
            # Envoi de l'email
            send_mail(
                subject=f"Invitation à rejoindre {context['site_name']}",
                message=plain_message,
                from_email=from_email,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            return True
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur lors de l'envoi de l'email d'invitation à {user.email}: {e}")
            return False
    
    @classmethod
    def send_password_reset_email(
        cls,
        user: User,
        username: str,
        password: str,
        reset_by: User
    ) -> bool:
        """
        Envoie l'email de réinitialisation de mot de passe.
        
        Args:
            user: L'utilisateur concerné
            username: Nom d'utilisateur
            password: Nouveau mot de passe temporaire
            reset_by: Utilisateur qui a réinitialisé le mot de passe
        
        Returns:
            True si l'email a été envoyé, False sinon
        """
        try:
            context = {
                'user': user,
                'username': username,
                'password': password,
                'reset_by': reset_by,
                'login_url': f"{getattr(settings, 'SITE_URL', '')}/accounts/login/",
                'site_name': getattr(settings, 'SITE_NAME', 'EEBC'),
            }
            
            # Rendu du template email
            html_message = render_to_string('accounts/emails/password_reset.html', context)
            plain_message = strip_tags(html_message)
            
            # Envoi de l'email
            send_mail(
                subject=f"Réinitialisation de votre mot de passe - {context['site_name']}",
                message=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            return True
            
        except Exception as e:
            print(f"Erreur lors de l'envoi de l'email de réinitialisation: {e}")
            return False
    
    @classmethod
    def resend_invitation(cls, user: User, reset_by: User) -> ServiceResult:
        """
        Renvoie l'email d'invitation à un utilisateur.
        
        Args:
            user: L'utilisateur à qui renvoyer l'invitation
            reset_by: L'utilisateur qui effectue l'action
        
        Returns:
            ServiceResult avec le nouveau mot de passe si succès
        """
        # Vérifier que l'utilisateur n'a pas encore activé son compte
        if not user.must_change_password:
            return ServiceResult.fail("L'utilisateur a déjà activé son compte")
        
        # Générer un nouveau mot de passe temporaire
        new_password = cls.generate_password()
        
        # Mettre à jour le mot de passe temporaire
        user.set_password(new_password)
        user.must_change_password = True
        user.save()
        
        # Renvoyer l'email
        if cls.send_invitation_email(user, user.username, new_password, reset_by):
            return ServiceResult.ok({'password': new_password})
        else:
            return ServiceResult.fail("Erreur lors de l'envoi de l'email")
    
    @classmethod
    def reset_user_password(cls, user: User, reset_by: User) -> ServiceResult:
        """
        Réinitialise le mot de passe d'un utilisateur actif.
        
        Args:
            user: L'utilisateur dont le mot de passe doit être réinitialisé
            reset_by: L'utilisateur qui effectue l'action
        
        Returns:
            ServiceResult avec le nouveau mot de passe si succès
        """
        # Générer un nouveau mot de passe temporaire
        new_password = cls.generate_password()
        
        # Réinitialiser le mot de passe et forcer le changement
        user.set_password(new_password)
        user.must_change_password = True
        user.save()
        
        # Envoyer l'email de réinitialisation
        if cls.send_password_reset_email(user, user.username, new_password, reset_by):
            return ServiceResult.ok({'password': new_password})
        else:
            return ServiceResult.fail("Erreur lors de l'envoi de l'email")

    # =========================================================================
    # IMPORT EN MASSE D'UTILISATEURS DEPUIS UN FICHIER EXCEL/CSV
    # =========================================================================
    EXPECTED_COLUMNS = {
        'first_name': ['first_name', 'prenom', 'prénom', 'firstname', 'first name'],
        'last_name': ['last_name', 'nom', 'lastname', 'last name', 'nom de famille'],
        'email': ['email', 'mail', 'e-mail', 'adresse email', 'adresse mail'],
        'phone': ['phone', 'telephone', 'téléphone', 'tel', 'tél', 'mobile'],
        'roles': ['role', 'roles', 'rôle', 'rôles', 'fonction'],
    }

    @classmethod
    def _normalize_header(cls, value) -> str:
        if value is None:
            return ''
        text = unicodedata.normalize('NFD', str(value).strip().lower())
        text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
        return text

    @classmethod
    def _map_columns(cls, headers: list) -> dict:
        """Associe les en-têtes du fichier aux champs internes."""
        normalized = [cls._normalize_header(h) for h in headers]
        mapping = {}
        for field, aliases in cls.EXPECTED_COLUMNS.items():
            for alias in aliases:
                if alias in normalized:
                    mapping[field] = normalized.index(alias)
                    break
        return mapping

    @classmethod
    def parse_users_file(cls, uploaded_file) -> ServiceResult:
        """
        Lit un fichier Excel/CSV et retourne une liste de lignes normalisées.
        Colonnes acceptées : first_name/prenom, last_name/nom, email, phone (optionnel), roles (optionnel).
        """
        import os

        name = getattr(uploaded_file, 'name', '') or ''
        ext = os.path.splitext(name)[1].lower()

        rows = []
        try:
            if ext in ('.xlsx', '.xls', '.xlsm'):
                from openpyxl import load_workbook
                uploaded_file.seek(0)
                wb = load_workbook(uploaded_file, data_only=True, read_only=True)
                ws = wb.active
                iterator = ws.iter_rows(values_only=True)
                try:
                    headers = list(next(iterator))
                except StopIteration:
                    return ServiceResult.fail("Fichier vide.")
                for row in iterator:
                    if row is None:
                        continue
                    if all((c is None or str(c).strip() == '') for c in row):
                        continue
                    rows.append(list(row))
            elif ext == '.csv':
                import csv, io
                uploaded_file.seek(0)
                raw = uploaded_file.read()
                if isinstance(raw, bytes):
                    try:
                        text = raw.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        text = raw.decode('latin-1')
                else:
                    text = raw
                # Détection simple du séparateur
                sample = text.splitlines()[0] if text else ''
                delimiter = ';' if sample.count(';') > sample.count(',') else ','
                reader = csv.reader(io.StringIO(text), delimiter=delimiter)
                try:
                    headers = next(reader)
                except StopIteration:
                    return ServiceResult.fail("Fichier vide.")
                for row in reader:
                    if not any((c or '').strip() for c in row):
                        continue
                    rows.append(row)
            else:
                return ServiceResult.fail(
                    "Format non supporté. Utilisez un fichier .xlsx, .xls ou .csv."
                )
        except Exception as e:
            return ServiceResult.fail(f"Impossible de lire le fichier : {e}")

        mapping = cls._map_columns(headers)

        missing = [k for k in ('first_name', 'last_name', 'email') if k not in mapping]
        if missing:
            return ServiceResult.fail(
                "Colonnes obligatoires manquantes : "
                + ", ".join(missing)
                + ". Colonnes détectées : "
                + ", ".join(str(h) for h in headers)
            )

        parsed = []
        for idx, row in enumerate(rows, start=2):  # ligne 1 = entêtes
            def _get(field):
                col = mapping.get(field)
                if col is None or col >= len(row):
                    return ''
                value = row[col]
                return '' if value is None else str(value).strip()

            entry = {
                'line': idx,
                'first_name': _get('first_name'),
                'last_name': _get('last_name'),
                'email': _get('email').lower(),
                'phone': _get('phone'),
                'roles_raw': _get('roles'),
            }
            parsed.append(entry)

        return ServiceResult.ok({'rows': parsed, 'headers': headers})

    @classmethod
    def bulk_create_users_from_file(
        cls,
        uploaded_file,
        created_by: User,
        default_roles: list = None,
        send_email: bool = True,
    ) -> ServiceResult:
        """
        Importe en masse des utilisateurs depuis un fichier Excel/CSV.
        Retourne un résumé : created, skipped, errors.
        """
        parse_result = cls.parse_users_file(uploaded_file)
        if not parse_result.success:
            return parse_result

        rows = parse_result.data['rows']
        if not rows:
            return ServiceResult.fail("Aucune ligne de données trouvée dans le fichier.")

        valid_role_codes = {choice[0] for choice in User.Role.choices}
        default_roles = [r for r in (default_roles or []) if r in valid_role_codes] or [User.Role.MEMBRE]

        created = []
        skipped = []
        errors = []
        seen_emails = set()

        for entry in rows:
            line = entry['line']
            first_name = entry['first_name']
            last_name = entry['last_name']
            email = entry['email']

            if not first_name or not last_name or not email:
                errors.append({
                    'line': line,
                    'email': email,
                    'reason': "Prénom, nom et email sont obligatoires.",
                })
                continue

            if '@' not in email or '.' not in email.split('@')[-1]:
                errors.append({'line': line, 'email': email, 'reason': "Email invalide."})
                continue

            if email in seen_emails:
                skipped.append({'line': line, 'email': email, 'reason': "Email dupliqué dans le fichier."})
                continue
            seen_emails.add(email)

            if User.objects.filter(email__iexact=email).exists():
                skipped.append({'line': line, 'email': email, 'reason': "Email déjà utilisé."})
                continue

            # Rôles : depuis la ligne (séparateur , ou ;) sinon défaut
            row_roles_raw = entry.get('roles_raw') or ''
            row_roles = [
                r.strip().lower()
                for r in row_roles_raw.replace(';', ',').split(',')
                if r.strip()
            ]
            row_roles = [r for r in row_roles if r in valid_role_codes]
            roles = row_roles or list(default_roles)

            result = cls.create_user_by_team(
                first_name=first_name.title(),
                last_name=last_name.upper(),
                email=email,
                roles=roles,
                created_by=created_by,
                phone=entry.get('phone', ''),
                send_email=send_email,
            )

            if result.success:
                created.append({
                    'line': line,
                    'email': email,
                    'username': result.data['username'],
                    'password': result.data['password'],
                })
            else:
                errors.append({'line': line, 'email': email, 'reason': result.error})

        return ServiceResult.ok({
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'total': len(rows),
        })
