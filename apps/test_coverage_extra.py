"""
Tests supplémentaires pour atteindre 60% de couverture.
Cible les vues worship, finance, events, communication, imports restantes.
"""
import io
import pytest
from datetime import date, time
from decimal import Decimal
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

User = get_user_model()


# =============================================================================
# FIXTURES worship
# =============================================================================

@pytest.fixture
def worship_service(db, event, admin_user):
    from apps.worship.models import WorshipService
    return WorshipService.objects.create(
        event=event,
        service_type='culte_dominical',
        created_by=admin_user,
    )


@pytest.fixture
def monthly_schedule(db, site_cayenne):
    from apps.worship.models import MonthlySchedule
    return MonthlySchedule.objects.create(
        month=6, year=2026, site=site_cayenne,
    )


@pytest.fixture
def service_template(db):
    from apps.worship.models import ServiceTemplate
    return ServiceTemplate.objects.create(
        name='Culte standard', service_type='culte_dominical',
        estimated_duration=90,
    )


# =============================================================================
# FIXTURES finance
# =============================================================================

@pytest.fixture
def finance_category(db):
    from apps.finance.models import FinanceCategory
    return FinanceCategory.objects.create(name='Dons', is_income=True)


@pytest.fixture
def budget_category_f(db):
    from apps.finance.models import BudgetCategory
    return BudgetCategory.objects.create(name='Fonctionnement')


# =============================================================================
# WORSHIP — vues détaillées
# =============================================================================

@pytest.mark.django_db
class TestWorshipDetailedViews:

    def test_service_detail(self, authenticated_client, worship_service):
        r = authenticated_client.get(
            reverse('worship:service_detail', args=[worship_service.pk])
        )
        assert r.status_code == 200

    def test_service_edit_get(self, authenticated_client, worship_service):
        r = authenticated_client.get(
            reverse('worship:service_edit', args=[worship_service.pk])
        )
        assert r.status_code == 200

    def test_service_edit_post(self, authenticated_client, worship_service, site_cayenne):
        r = authenticated_client.post(
            reverse('worship:service_edit', args=[worship_service.pk]),
            {'title': 'Culte modifié', 'date': '2026-06-07',
             'start_time': '10:00', 'site': site_cayenne.pk},
        )
        assert r.status_code in (200, 302)

    def test_service_delete(self, authenticated_client, worship_service):
        r = authenticated_client.post(
            reverse('worship:service_delete', args=[worship_service.pk])
        )
        assert r.status_code in (200, 302)

    def test_schedule_detail(self, authenticated_client, monthly_schedule):
        r = authenticated_client.get(
            reverse('worship:schedule_detail', args=[monthly_schedule.pk])
        )
        assert r.status_code == 200

    def test_schedule_edit_get(self, authenticated_client, monthly_schedule):
        r = authenticated_client.get(
            reverse('worship:schedule_edit', args=[monthly_schedule.pk])
        )
        assert r.status_code == 200

    def test_generate_sundays(self, authenticated_client, monthly_schedule):
        r = authenticated_client.post(
            reverse('worship:generate_sundays', args=[monthly_schedule.pk])
        )
        assert r.status_code in (200, 302)

    def test_publish_schedule(self, authenticated_client, monthly_schedule):
        r = authenticated_client.post(
            reverse('worship:publish_schedule', args=[monthly_schedule.pk])
        )
        assert r.status_code in (200, 302)

    def test_template_detail(self, authenticated_client, service_template):
        r = authenticated_client.get(
            reverse('worship:template_detail', args=[service_template.pk])
        )
        assert r.status_code == 200

    def test_template_update_get(self, authenticated_client, service_template):
        r = authenticated_client.get(
            reverse('worship:template_update', args=[service_template.pk])
        )
        assert r.status_code == 200

    def test_template_update_post(self, authenticated_client, service_template):
        r = authenticated_client.post(
            reverse('worship:template_update', args=[service_template.pk]),
            {'name': 'Modifié', 'service_type': 'culte_dominical',
             'estimated_duration': 120},
        )
        assert r.status_code in (200, 302)

    def test_template_delete(self, authenticated_client, service_template):
        r = authenticated_client.post(
            reverse('worship:template_delete', args=[service_template.pk])
        )
        assert r.status_code in (200, 302)

    def test_template_item_create_get(self, authenticated_client, service_template):
        r = authenticated_client.get(
            reverse('worship:template_item_create', args=[service_template.pk])
        )
        assert r.status_code == 200

    def test_template_item_create_post(self, authenticated_client, service_template):
        r = authenticated_client.post(
            reverse('worship:template_item_create', args=[service_template.pk]),
            {'title': 'Louange', 'duration': 15, 'order': 1},
        )
        assert r.status_code in (200, 302)

    def test_plan_edit(self, authenticated_client, worship_service):
        r = authenticated_client.get(
            reverse('worship:plan_edit', args=[worship_service.pk])
        )
        assert r.status_code == 200

    def test_plan_item_add(self, authenticated_client, worship_service):
        r = authenticated_client.post(
            reverse('worship:plan_item_add', args=[worship_service.pk]),
            {'title': 'Prière', 'duration': 10, 'order': 1},
        )
        assert r.status_code in (200, 302)

    def test_run_sheet_pdf(self, authenticated_client, worship_service):
        r = authenticated_client.get(
            reverse('worship:run_sheet_pdf', args=[worship_service.pk])
        )
        assert r.status_code == 200




# =============================================================================
# FINANCE — vues détaillées
# =============================================================================

@pytest.mark.django_db
class TestFinanceDetailedViews:

    def test_budget_delete_get(self, authenticated_client, admin_user):
        from apps.finance.models import Budget

        budget = Budget.objects.create(
            name='Budget suppression',
            year=2026,
            total_requested=Decimal('2500.00'),
            created_by=admin_user,
            status='draft',
        )

        r = authenticated_client.get(reverse('finance:budget_delete', args=[budget.pk]))
        assert r.status_code == 200

    def test_budget_delete_post(self, authenticated_client, admin_user):
        from apps.finance.models import Budget

        budget = Budget.objects.create(
            name='Budget suppression POST',
            year=2026,
            total_requested=Decimal('3000.00'),
            created_by=admin_user,
            status='draft',
        )

        r = authenticated_client.post(reverse('finance:budget_delete', args=[budget.pk]))
        assert r.status_code == 302
        assert not Budget.objects.filter(pk=budget.pk).exists()

    def test_budget_delete_blocked_when_transactions_exist(
        self,
        authenticated_client,
        admin_user,
        site_cayenne,
        finance_category,
        budget_category_f,
    ):
        from apps.finance.models import Budget, BudgetItem, FinancialTransaction

        budget = Budget.objects.create(
            name='Budget verrouillé',
            year=2026,
            total_requested=Decimal('1500.00'),
            created_by=admin_user,
            status='draft',
        )
        item = BudgetItem.objects.create(
            budget=budget,
            category=budget_category_f,
            requested_amount=Decimal('1500.00'),
            description='Ligne test',
            justification='Test suppression',
        )
        FinancialTransaction.objects.create(
            site=site_cayenne,
            amount=Decimal('150.00'),
            transaction_type='depense',
            payment_method='especes',
            status='valide',
            transaction_date=date(2026, 4, 1),
            category=finance_category,
            budget_item=item,
            recorded_by=admin_user,
        )

        r = authenticated_client.post(reverse('finance:budget_delete', args=[budget.pk]))
        assert r.status_code == 302
        assert Budget.objects.filter(pk=budget.pk).exists()

    def test_transaction_list(self, authenticated_client, financial_transaction):
        r = authenticated_client.get(reverse('finance:transaction_list'))
        assert r.status_code == 200

    def test_transaction_detail(self, authenticated_client, financial_transaction):
        r = authenticated_client.get(
            reverse('finance:transaction_detail', args=[financial_transaction.pk])
        )
        assert r.status_code == 200

    def test_tax_receipt_list(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:tax_receipt_list'))
        assert r.status_code == 200

    def test_budget_category_list(self, authenticated_client, budget_category_f):
        r = authenticated_client.get(reverse('finance:budget_category_list'))
        assert r.status_code == 200

    def test_budget_category_create_get(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:budget_category_create'))
        assert r.status_code == 200

    def test_budget_category_create_post(self, authenticated_client):
        r = authenticated_client.post(reverse('finance:budget_category_create'), {
            'name': 'Missions', 'color': '#00FF00',
        })
        assert r.status_code in (200, 302)

    def test_budget_category_update_get(self, authenticated_client, budget_category_f):
        r = authenticated_client.get(
            reverse('finance:budget_category_update', args=[budget_category_f.pk])
        )
        assert r.status_code == 200

    def test_budget_category_delete(self, authenticated_client, budget_category_f):
        r = authenticated_client.post(
            reverse('finance:budget_category_delete', args=[budget_category_f.pk])
        )
        assert r.status_code in (200, 302)

    def test_finance_category_list(self, authenticated_client, finance_category):
        r = authenticated_client.get(reverse('finance:finance_category_list'))
        assert r.status_code == 200

    def test_finance_category_create_get(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:finance_category_create'))
        assert r.status_code == 200

    def test_finance_category_create_post(self, authenticated_client):
        r = authenticated_client.post(reverse('finance:finance_category_create'), {
            'name': 'Offrandes', 'is_income': True,
        })
        assert r.status_code in (200, 302)

    def test_finance_category_update_get(self, authenticated_client, finance_category):
        r = authenticated_client.get(
            reverse('finance:finance_category_update', args=[finance_category.pk])
        )
        assert r.status_code == 200

    def test_finance_category_delete(self, authenticated_client, finance_category):
        r = authenticated_client.post(
            reverse('finance:finance_category_delete', args=[finance_category.pk])
        )
        assert r.status_code in (200, 302)

    def test_receipt_proof_list(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:receipt_proof_list'))
        assert r.status_code == 200

    def test_finance_import_excel_get(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:import_excel'))
        assert r.status_code == 200

    def test_finance_import_excel_template(self, authenticated_client):
        r = authenticated_client.get(reverse('finance:import_excel_template'))
        assert r.status_code == 200
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in r['Content-Type']

    def test_finance_import_excel_post(self, authenticated_client, site_cayenne):
        from apps.finance.models import FinanceCategory, FinancialTransaction

        workbook = Workbook()
        workbook.remove(workbook.active)

        categories = workbook.create_sheet('finance_categories')
        categories.append(['name', 'is_income', 'description'])
        categories.append(['Offrandes importées', True, 'Import test'])

        transactions = workbook.create_sheet('transactions')
        transactions.append([
            'reference', 'site_name', 'amount', 'transaction_type', 'payment_method', 'status', 'transaction_date',
            'description', 'category_name',
        ])
        transactions.append([
            'IMP-TRX-001', site_cayenne.name, 275.50, 'offrande', 'especes', 'valide', date(2026, 4, 17),
            'Import unitaire', 'Offrandes importées',
        ])

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        uploaded_file = SimpleUploadedFile(
            'finance_import_test.xlsx',
            payload.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        r = authenticated_client.post(
            reverse('finance:import_excel'),
            {'file': uploaded_file},
        )

        assert r.status_code == 302
        assert FinanceCategory.objects.filter(name='Offrandes importées').exists()
        assert FinancialTransaction.objects.filter(reference='IMP-TRX-001').exists()


# =============================================================================
# EVENTS — vues restantes
# =============================================================================

@pytest.mark.django_db
class TestEventsDetailedViews:

    def test_event_detail(self, authenticated_client, event):
        r = authenticated_client.get(reverse('events:detail', args=[event.pk]))
        assert r.status_code == 200

    def test_events_json(self, authenticated_client, event):
        r = authenticated_client.get(reverse('events:events_json'))
        assert r.status_code == 200

    def test_calendar_pdf(self, authenticated_client):
        r = authenticated_client.get(reverse('events:calendar_pdf'))
        assert r.status_code == 200


# =============================================================================
# COMMUNICATION — vues restantes
# =============================================================================

@pytest.mark.django_db
class TestCommunicationDetailedViews:

    def test_notifications_list(self, authenticated_client, admin_user):
        from apps.communication.models import Notification
        Notification.objects.create(user=admin_user, title='N1', message='m1')
        r = authenticated_client.get(reverse('communication:notifications'))
        assert r.status_code == 200

    def test_announcements_list(self, authenticated_client, admin_user):
        from apps.communication.models import Announcement
        Announcement.objects.create(
            title='Ann1', content='Content', created_by=admin_user
        )
        r = authenticated_client.get(reverse('communication:announcements'))
        assert r.status_code == 200

    def test_announcement_detail(self, authenticated_client, admin_user):
        from apps.communication.models import Announcement
        ann = Announcement.objects.create(
            title='Detail', content='Content', created_by=admin_user
        )
        r = authenticated_client.get(
            reverse('communication:announcement_detail', args=[ann.pk])
        )
        assert r.status_code == 200

    def test_email_logs(self, authenticated_client):
        r = authenticated_client.get(reverse('communication:email_logs'))
        assert r.status_code == 200

    def test_email_logs_management(self, authenticated_client):
        r = authenticated_client.get(reverse('communication:email_logs_management'))
        assert r.status_code == 200

    def test_email_logs_clear_old(self, authenticated_client):
        r = authenticated_client.post(reverse('communication:email_logs_clear_old'))
        assert r.status_code in (200, 302)

    def test_sms_logs(self, authenticated_client):
        r = authenticated_client.get(reverse('communication:sms_logs'))
        assert r.status_code == 200


# =============================================================================
# IMPORTS — exports restants
# =============================================================================

@pytest.mark.django_db
class TestImportsExportViews:

    def test_export_groups(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:export_groups'))
        assert r.status_code == 200

    def test_export_inventory(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:export_inventory'))
        assert r.status_code == 200

    def test_export_transport(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:export_transport'))
        assert r.status_code == 200

    def test_export_communication(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:export_communication'))
        assert r.status_code == 200

    def test_download_template_members(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:template', args=['members']))
        assert r.status_code == 200

    def test_download_template_children(self, authenticated_client):
        r = authenticated_client.get(reverse('imports:template', args=['children']))
        assert r.status_code == 200


# =============================================================================
# MEMBERS — vues restantes
# =============================================================================

@pytest.mark.django_db
class TestMembersDetailedViews:

    def test_member_list(self, authenticated_client, member):
        r = authenticated_client.get(reverse('members:list'))
        assert r.status_code == 200

    def test_member_detail(self, authenticated_client, member):
        r = authenticated_client.get(reverse('members:detail', args=[member.pk]))
        assert r.status_code == 200


# =============================================================================
# EXPORTS — core export views (Excel)
# =============================================================================

@pytest.mark.django_db
class TestExportViews:

    def test_children_excel(self, authenticated_client):
        r = authenticated_client.get(reverse('exports:children_excel'))
        assert r.status_code == 200

    def test_members_excel(self, authenticated_client, member):
        r = authenticated_client.get(reverse('exports:members_excel'))
        assert r.status_code == 200

    def test_events_excel(self, authenticated_client, event):
        r = authenticated_client.get(reverse('exports:events_excel'))
        assert r.status_code == 200

    def test_groups_excel(self, authenticated_client):
        from apps.groups.models import Group
        Group.objects.create(name='Export test')
        r = authenticated_client.get(reverse('exports:groups_excel'))
        assert r.status_code == 200

    def test_budgets_excel(self, authenticated_client):
        r = authenticated_client.get(reverse('exports:budgets_excel'))
        assert r.status_code == 200

    def test_transactions_excel(self, authenticated_client, financial_transaction):
        r = authenticated_client.get(reverse('exports:transactions_excel'))
        assert r.status_code == 200

    def test_departments_excel(self, authenticated_client):
        r = authenticated_client.get(reverse('exports:departments_excel'))
        assert r.status_code == 200

    def test_drivers_excel(self, authenticated_client):
        r = authenticated_client.get(reverse('exports:drivers_excel'))
        assert r.status_code == 200

    def test_attendance_excel(self, authenticated_client):
        r = authenticated_client.get(reverse('exports:attendance_excel'))
        assert r.status_code == 200
