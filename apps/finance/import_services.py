from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
import unicodedata

from django.db import transaction
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from apps.accounts.models import User
from apps.core.models import Site
from apps.departments.models import Department
from apps.events.models import Event
from apps.finance.models import (
    Budget,
    BudgetCategory,
    BudgetForecast,
    BudgetItem,
    BudgetLine,
    BudgetRequest,
    FinanceCategory,
    FinancialTransaction,
    ForecastLine,
)
from apps.groups.models import Group
from apps.members.models import Member


class FinanceImportError(ValueError):
    pass


SECTION_CHOICES = ['categories', 'budget_lines', 'budgets', 'forecasts', 'transactions']

FINANCE_IMPORT_SHEET_SPECS = [
    {
        'name': 'finance_categories',
        'label': 'Categories financieres',
        'section': 'categories',
        'description': 'Cree ou met a jour les categories de transactions.',
        'required_columns': ['name', 'is_income'],
        'columns': ['name', 'description', 'is_income', 'budget_annual', 'parent_name', 'is_active'],
    },
    {
        'name': 'budget_categories',
        'label': 'Categories de budget',
        'section': 'categories',
        'description': 'Cree ou met a jour les categories de budget.',
        'required_columns': ['name'],
        'columns': ['name', 'description', 'color', 'is_active'],
    },
    {
        'name': 'budget_lines',
        'label': 'Lignes budgetaires',
        'section': 'budget_lines',
        'description': 'Importe les enveloppes budgetaires annuelles ou mensuelles.',
        'required_columns': ['category_name', 'year', 'planned_amount'],
        'columns': ['category_name', 'year', 'month', 'planned_amount', 'notes'],
    },
    {
        'name': 'budgets',
        'label': 'Budgets',
        'section': 'budgets',
        'description': 'Importe les entetes de budgets groupes/departements.',
        'required_columns': ['name', 'year'],
        'columns': [
            'name', 'year', 'group_name', 'group_site_name', 'department_name', 'department_site_name',
            'total_requested', 'total_approved', 'status', 'description', 'approval_notes',
            'created_by', 'approved_by', 'submitted_at', 'approved_at',
        ],
    },
    {
        'name': 'budget_items',
        'label': 'Lignes de budgets',
        'section': 'budgets',
        'description': 'Associe des lignes de budget a un budget existant dans la feuille budgets.',
        'required_columns': ['budget_name', 'budget_year', 'category_name', 'requested_amount'],
        'columns': [
            'budget_name', 'budget_year', 'group_name', 'group_site_name', 'department_name', 'department_site_name',
            'category_name', 'requested_amount', 'approved_amount', 'approval_status', 'approved_by', 'approved_at',
            'approval_comments', 'rejection_reason', 'description', 'justification', 'priority',
        ],
    },
    {
        'name': 'budget_requests',
        'label': 'Demandes de budget',
        'section': 'budgets',
        'description': 'Importe les demandes de budget ponctuelles.',
        'required_columns': ['title', 'requested_amount', 'category_name'],
        'columns': [
            'title', 'description', 'requested_amount', 'approved_amount', 'category_name',
            'group_name', 'group_site_name', 'department_name', 'department_site_name',
            'urgency', 'needed_by', 'status', 'requested_by', 'reviewed_by', 'reviewed_at',
            'justification', 'review_notes',
        ],
    },
    {
        'name': 'forecasts',
        'label': 'Previsionnels',
        'section': 'forecasts',
        'description': 'Importe les entetes de previsionnels.',
        'required_columns': ['year', 'scenario', 'name'],
        'columns': ['name', 'year', 'scenario', 'description', 'is_active', 'created_by'],
    },
    {
        'name': 'forecast_lines',
        'label': 'Lignes de previsionnels',
        'section': 'forecasts',
        'description': 'Associe des lignes mensuelles a un previsionnel existant dans la feuille forecasts.',
        'required_columns': ['forecast_year', 'forecast_scenario', 'label', 'line_type'],
        'columns': [
            'forecast_year', 'forecast_scenario', 'label', 'line_type', 'category_name',
            'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec', 'notes',
        ],
    },
    {
        'name': 'transactions',
        'label': 'Transactions',
        'section': 'transactions',
        'description': 'Importe les transactions financieres et leurs rattachements.',
        'required_columns': ['reference', 'amount', 'transaction_type', 'transaction_date'],
        'columns': [
            'reference', 'site_name', 'amount', 'transaction_type', 'payment_method', 'status', 'transaction_date',
            'description', 'category_name', 'member_id', 'event_title', 'event_start_date', 'event_site_name',
            'budget_name', 'budget_year', 'budget_group_name', 'budget_group_site_name',
            'budget_department_name', 'budget_department_site_name', 'budget_category_name',
            'recorded_by', 'validated_by', 'validated_at', 'notes', 'is_deleted', 'deleted_at', 'deleted_by',
        ],
    },
]

SHEET_SPECS_BY_NAME = {spec['name']: spec for spec in FINANCE_IMPORT_SHEET_SPECS}
DATETIME_FIELDS = {'submitted_at', 'approved_at', 'reviewed_at', 'validated_at', 'deleted_at'}
DATE_FIELDS = {'needed_by', 'transaction_date', 'event_start_date'}
INTEGER_FIELDS = {'year', 'month', 'budget_year', 'forecast_year', 'priority'}
BOOLEAN_FIELDS = {'is_income', 'is_active', 'is_deleted'}


def _normalize_identifier(value):
    if value is None:
        return ''
    value = str(value).strip().lower()
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    for char in ('/', '-', '.', ':'):
        value = value.replace(char, ' ')
    return '_'.join(part for part in value.split() if part)


def _is_empty(value):
    return value is None or (isinstance(value, str) and not value.strip())


def build_finance_import_template_workbook():
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = 'instructions'
    instructions.append(['sheet_name', 'section', 'description', 'required_columns', 'all_columns'])
    for cell in instructions[1]:
        cell.font = Font(bold=True)

    for spec in FINANCE_IMPORT_SHEET_SPECS:
        instructions.append([
            spec['name'],
            spec['section'],
            spec['description'],
            ', '.join(spec['required_columns']),
            ', '.join(spec['columns']),
        ])

        sheet = workbook.create_sheet(title=spec['name'])
        sheet.append(spec['columns'])
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    instructions.column_dimensions['A'].width = 24
    instructions.column_dimensions['B'].width = 16
    instructions.column_dimensions['C'].width = 68
    instructions.column_dimensions['D'].width = 48
    instructions.column_dimensions['E'].width = 92
    return workbook


class FinanceExcelWorkbookParser:
    supported_extensions = {'.xlsx', '.xlsm'}

    def parse(self, workbook_source):
        workbook = self._load_workbook(workbook_source)
        sheet_records = {}

        for worksheet in workbook.worksheets:
            normalized_name = _normalize_identifier(worksheet.title)
            if normalized_name not in SHEET_SPECS_BY_NAME:
                continue
            records = self._read_sheet(worksheet, SHEET_SPECS_BY_NAME[normalized_name])
            sheet_records[normalized_name] = records

        if not sheet_records:
            raise FinanceImportError(
                "Aucune feuille reconnue. Utilisez le modele fourni et conservez les noms d'onglets attendus."
            )

        self._validate_dependencies(sheet_records)

        sections = []
        bundle = {
            'meta': {
                'source': 'excel',
                'sections': sections,
                'schema_version': 1,
            }
        }

        if 'finance_categories' in sheet_records or 'budget_categories' in sheet_records:
            sections.append('categories')
            bundle['finance_categories'] = sheet_records.get('finance_categories', [])
            bundle['budget_categories'] = sheet_records.get('budget_categories', [])

        if 'budget_lines' in sheet_records:
            sections.append('budget_lines')
            bundle['budget_lines'] = sheet_records['budget_lines']

        budget_related_sheets = {'budgets', 'budget_items', 'budget_requests'}
        if budget_related_sheets & set(sheet_records):
            sections.append('budgets')
            bundle['budgets'] = self._build_budgets(sheet_records)
            bundle['budget_requests'] = self._build_budget_requests(sheet_records.get('budget_requests', []))

        forecast_related_sheets = {'forecasts', 'forecast_lines'}
        if forecast_related_sheets & set(sheet_records):
            sections.append('forecasts')
            bundle['forecasts'] = self._build_forecasts(sheet_records)

        if 'transactions' in sheet_records:
            sections.append('transactions')
            bundle['transactions'] = self._build_transactions(sheet_records['transactions'])

        return bundle, sections

    def _load_workbook(self, workbook_source):
        if hasattr(workbook_source, 'name'):
            extension = Path(workbook_source.name).suffix.lower()
            if extension and extension not in self.supported_extensions:
                raise FinanceImportError(
                    f"Format non pris en charge: {extension}. Utilisez un fichier .xlsx ou .xlsm."
                )
        if hasattr(workbook_source, 'seek'):
            workbook_source.seek(0)
        return load_workbook(workbook_source, data_only=True, read_only=True)

    def _read_sheet(self, worksheet, spec):
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = rows[0]
        headers = [_normalize_identifier(value) for value in raw_headers]
        if not any(headers):
            return []

        missing = [column for column in spec['required_columns'] if column not in headers]
        if missing:
            raise FinanceImportError(
                f"La feuille {spec['name']} ne contient pas les colonnes obligatoires: {', '.join(missing)}."
            )

        records = []
        for index, row in enumerate(rows[1:], start=2):
            if all(_is_empty(value) for value in row):
                continue

            record = {}
            for header, value in zip(headers, row):
                if not header:
                    continue
                record[header] = self._normalize_cell(value, header)
            missing_values = [column for column in spec['required_columns'] if _is_empty(record.get(column))]
            if missing_values:
                raise FinanceImportError(
                    f"La feuille {spec['name']} contient une ligne incomplète ({index}) : {', '.join(missing_values)}."
                )
            records.append(record)
        return records

    def _normalize_cell(self, value, header):
        if value is None:
            return None
        if isinstance(value, datetime):
            if header in DATE_FIELDS:
                return value.date().isoformat()
            return value.replace(microsecond=0).isoformat()
        if isinstance(value, date):
            if header in DATETIME_FIELDS:
                return datetime.combine(value, time.min).isoformat()
            return value.isoformat()
        if header in INTEGER_FIELDS and isinstance(value, float) and value.is_integer():
            return int(value)
        if header in BOOLEAN_FIELDS:
            return self._normalize_boolean(value)
        if isinstance(value, str):
            return value.strip()
        return value

    def _normalize_boolean(self, value):
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        normalized = str(value).strip().lower()
        return normalized in {'1', 'true', 'vrai', 'oui', 'yes', 'y', 'x'}

    def _validate_dependencies(self, sheet_records):
        if 'budget_items' in sheet_records and 'budgets' not in sheet_records:
            raise FinanceImportError("La feuille budget_items exige la presence de la feuille budgets.")
        if 'forecast_lines' in sheet_records and 'forecasts' not in sheet_records:
            raise FinanceImportError("La feuille forecast_lines exige la presence de la feuille forecasts.")

    def _build_budgets(self, sheet_records):
        budgets = {}
        for record in sheet_records.get('budgets', []):
            self._validate_single_entity(record, 'budgets')
            key = self._budget_key(
                record.get('name'),
                record.get('year'),
                record.get('group_name'),
                record.get('group_site_name'),
                record.get('department_name'),
                record.get('department_site_name'),
            )
            budgets[key] = {
                'name': record.get('name'),
                'year': record.get('year'),
                'group': self._entity_payload(record, 'group_name', 'group_site_name'),
                'department': self._entity_payload(record, 'department_name', 'department_site_name'),
                'total_requested': record.get('total_requested') or Decimal('0'),
                'total_approved': record.get('total_approved') or Decimal('0'),
                'status': record.get('status') or Budget.Status.DRAFT,
                'created_by': record.get('created_by'),
                'approved_by': record.get('approved_by'),
                'submitted_at': record.get('submitted_at'),
                'approved_at': record.get('approved_at'),
                'description': record.get('description') or '',
                'approval_notes': record.get('approval_notes') or '',
                'items': [],
            }

        for record in sheet_records.get('budget_items', []):
            self._validate_single_entity(record, 'budget_items')
            key = self._budget_key(
                record.get('budget_name'),
                record.get('budget_year'),
                record.get('group_name'),
                record.get('group_site_name'),
                record.get('department_name'),
                record.get('department_site_name'),
            )
            if key not in budgets:
                raise FinanceImportError(
                    f"Aucun budget correspondant pour la ligne de budget {record.get('category_name')} ({record.get('budget_name')})."
                )
            budgets[key]['items'].append({
                'category_name': record.get('category_name'),
                'requested_amount': record.get('requested_amount') or Decimal('0'),
                'approved_amount': record.get('approved_amount') or Decimal('0'),
                'approval_status': record.get('approval_status') or BudgetItem.ApprovalStatus.PENDING,
                'approved_by': record.get('approved_by'),
                'approved_at': record.get('approved_at'),
                'approval_comments': record.get('approval_comments') or '',
                'rejection_reason': record.get('rejection_reason') or '',
                'description': record.get('description') or '',
                'justification': record.get('justification') or '',
                'priority': record.get('priority') or 1,
            })

        for budget in budgets.values():
            if not budget['total_requested'] and budget['items']:
                budget['total_requested'] = sum(
                    Decimal(str(item['requested_amount'] or 0)) for item in budget['items']
                )
            if not budget['total_approved'] and budget['items']:
                budget['total_approved'] = sum(
                    Decimal(str(item['approved_amount'] or 0)) for item in budget['items']
                )

        return list(budgets.values())

    def _build_budget_requests(self, records):
        requests = []
        for record in records:
            self._validate_single_entity(record, 'budget_requests')
            requests.append({
                'title': record.get('title'),
                'description': record.get('description') or '',
                'requested_amount': record.get('requested_amount') or Decimal('0'),
                'approved_amount': record.get('approved_amount') or Decimal('0'),
                'category_name': record.get('category_name'),
                'group': self._entity_payload(record, 'group_name', 'group_site_name'),
                'department': self._entity_payload(record, 'department_name', 'department_site_name'),
                'urgency': record.get('urgency') or BudgetRequest.Urgency.MEDIUM,
                'needed_by': record.get('needed_by'),
                'status': record.get('status') or BudgetRequest.Status.PENDING,
                'requested_by': record.get('requested_by'),
                'reviewed_by': record.get('reviewed_by'),
                'reviewed_at': record.get('reviewed_at'),
                'justification': record.get('justification') or '',
                'review_notes': record.get('review_notes') or '',
            })
        return requests

    def _build_forecasts(self, sheet_records):
        forecasts = {}
        for record in sheet_records.get('forecasts', []):
            key = (record.get('year'), record.get('scenario'))
            forecasts[key] = {
                'name': record.get('name'),
                'year': record.get('year'),
                'scenario': record.get('scenario'),
                'description': record.get('description') or '',
                'is_active': True if record.get('is_active') is None else record.get('is_active'),
                'created_by': record.get('created_by'),
                'lines': [],
            }

        for record in sheet_records.get('forecast_lines', []):
            key = (record.get('forecast_year'), record.get('forecast_scenario'))
            if key not in forecasts:
                raise FinanceImportError(
                    f"Aucun previsionnel correspondant pour la ligne {record.get('label')} ({record.get('forecast_year')} / {record.get('forecast_scenario')})."
                )
            forecasts[key]['lines'].append({
                'label': record.get('label'),
                'line_type': record.get('line_type'),
                'category_name': record.get('category_name'),
                'jan': record.get('jan') or Decimal('0'),
                'feb': record.get('feb') or Decimal('0'),
                'mar': record.get('mar') or Decimal('0'),
                'apr': record.get('apr') or Decimal('0'),
                'may': record.get('may') or Decimal('0'),
                'jun': record.get('jun') or Decimal('0'),
                'jul': record.get('jul') or Decimal('0'),
                'aug': record.get('aug') or Decimal('0'),
                'sep': record.get('sep') or Decimal('0'),
                'oct': record.get('oct') or Decimal('0'),
                'nov': record.get('nov') or Decimal('0'),
                'dec': record.get('dec') or Decimal('0'),
                'notes': record.get('notes') or '',
            })
        return list(forecasts.values())

    def _build_transactions(self, records):
        transactions = []
        for record in records:
            transactions.append({
                'reference': record.get('reference'),
                'site_name': record.get('site_name'),
                'amount': record.get('amount'),
                'transaction_type': record.get('transaction_type'),
                'payment_method': record.get('payment_method') or FinancialTransaction.PaymentMethod.ESPECES,
                'status': record.get('status') or FinancialTransaction.Status.EN_ATTENTE,
                'transaction_date': record.get('transaction_date'),
                'description': record.get('description') or '',
                'category_name': record.get('category_name'),
                'member_id': record.get('member_id'),
                'event': self._event_payload(record),
                'budget_item': self._budget_item_payload(record),
                'recorded_by': record.get('recorded_by'),
                'validated_by': record.get('validated_by'),
                'validated_at': record.get('validated_at'),
                'notes': record.get('notes') or '',
                'is_deleted': record.get('is_deleted') or False,
                'deleted_at': record.get('deleted_at'),
                'deleted_by': record.get('deleted_by'),
            })
        return transactions

    def _budget_key(self, name, year, group_name, group_site_name, department_name, department_site_name):
        return (
            name,
            int(year) if year is not None else None,
            group_name,
            group_site_name,
            department_name,
            department_site_name,
        )

    def _entity_payload(self, record, name_key, site_key):
        if not record.get(name_key):
            return None
        return {
            'name': record.get(name_key),
            'site_name': record.get(site_key),
        }

    def _event_payload(self, record):
        has_event_data = any(record.get(key) for key in ('event_title', 'event_start_date', 'event_site_name'))
        if not has_event_data:
            return None
        if not record.get('event_title') or not record.get('event_start_date'):
            raise FinanceImportError(
                "Chaque transaction liee a un evenement doit fournir event_title et event_start_date."
            )
        return {
            'title': record.get('event_title'),
            'start_date': record.get('event_start_date'),
            'site_name': record.get('event_site_name'),
        }

    def _budget_item_payload(self, record):
        budget_keys = (
            'budget_name', 'budget_year', 'budget_group_name', 'budget_group_site_name',
            'budget_department_name', 'budget_department_site_name', 'budget_category_name',
        )
        if not any(record.get(key) for key in budget_keys):
            return None
        if not record.get('budget_name') or record.get('budget_year') is None or not record.get('budget_category_name'):
            raise FinanceImportError(
                "Chaque transaction liee a un budget doit fournir budget_name, budget_year et budget_category_name."
            )
        if record.get('budget_group_name') and record.get('budget_department_name'):
            raise FinanceImportError(
                "Une transaction ne peut pas pointer a la fois vers un groupe et un departement pour la meme ligne de budget."
            )
        return {
            'budget_name': record.get('budget_name'),
            'budget_year': record.get('budget_year'),
            'group': self._entity_payload(record, 'budget_group_name', 'budget_group_site_name'),
            'department': self._entity_payload(record, 'budget_department_name', 'budget_department_site_name'),
            'category_name': record.get('budget_category_name'),
        }

    def _validate_single_entity(self, record, sheet_name):
        if record.get('group_name') and record.get('department_name'):
            raise FinanceImportError(
                f"La feuille {sheet_name} ne peut pas definir group_name et department_name sur la meme ligne."
            )


class FinanceBundleImporter:
    SECTION_CHOICES = SECTION_CHOICES

    def __init__(self):
        self.stats = {}
        self.warned = set()
        self.warnings = []

    def import_bundle(self, bundle, sections=None, dry_run=False):
        selected_sections = sections or bundle.get('meta', {}).get('sections') or self.SECTION_CHOICES
        self.stats = {}
        self.warned = set()
        self.warnings = []

        with transaction.atomic():
            if 'categories' in selected_sections:
                self.import_finance_categories(bundle.get('finance_categories', []))
                self.import_budget_categories(bundle.get('budget_categories', []))

            if 'budget_lines' in selected_sections:
                self.import_budget_lines(bundle.get('budget_lines', []))

            if 'budgets' in selected_sections:
                self.import_budgets(bundle.get('budgets', []))
                self.import_budget_requests(bundle.get('budget_requests', []))

            if 'forecasts' in selected_sections:
                self.import_forecasts(bundle.get('forecasts', []))

            if 'transactions' in selected_sections:
                self.import_transactions(bundle.get('transactions', []))

            if dry_run:
                transaction.set_rollback(True)

        return {
            'stats': self.stats,
            'warnings': self.warnings,
        }

    def bump(self, model_name, action):
        self.stats.setdefault(model_name, {'created': 0, 'updated': 0})
        self.stats[model_name][action] += 1

    def warn_once(self, key, message):
        if key in self.warned:
            return
        self.warned.add(key)
        self.warnings.append(message)

    def to_decimal(self, value):
        if value in (None, ''):
            return None
        return Decimal(str(value))

    def to_date(self, value):
        if not value:
            return None
        return parse_date(str(value))

    def to_datetime(self, value):
        if not value:
            return None
        return parse_datetime(str(value))

    def apply_values(self, obj, values):
        changed = False
        for field, value in values.items():
            if getattr(obj, field) != value:
                setattr(obj, field, value)
                changed = True
        if changed:
            obj.save()
        return changed

    def resolve_user(self, username):
        if not username:
            return None
        user = User.objects.filter(username=username).first()
        if user:
            return user
        self.warn_once(f'user:{username}', f'Utilisateur introuvable: {username}. Valeur ignoree.')
        return None

    def resolve_site(self, site_name):
        if not site_name:
            return None
        sites = list(Site.objects.filter(name=site_name))
        if len(sites) == 1:
            return sites[0]
        if not sites:
            self.warn_once(f'site:{site_name}', f'Site introuvable: {site_name}. Valeur ignoree.')
            return None
        self.warn_once(f'site_multi:{site_name}', f'Plusieurs sites nommes {site_name}. Premier enregistrement utilise.')
        return sites[0]

    def resolve_group(self, payload):
        if not payload:
            return None
        queryset = Group.objects.filter(name=payload['name'])
        site_name = payload.get('site_name')
        if site_name:
            queryset = queryset.filter(site__name=site_name)
        groups = list(queryset)
        if len(groups) == 1:
            return groups[0]
        if not groups:
            raise FinanceImportError(f"Groupe introuvable: {payload['name']}")
        raise FinanceImportError(f"Groupe ambigu: {payload['name']}")

    def resolve_department(self, payload):
        if not payload:
            return None
        queryset = Department.objects.filter(name=payload['name'])
        site_name = payload.get('site_name')
        if site_name:
            queryset = queryset.filter(site__name=site_name)
        departments = list(queryset)
        if len(departments) == 1:
            return departments[0]
        if not departments:
            raise FinanceImportError(f"Departement introuvable: {payload['name']}")
        raise FinanceImportError(f"Departement ambigu: {payload['name']}")

    def resolve_member(self, member_id):
        if not member_id:
            return None
        member = Member.objects.filter(member_id=member_id).first()
        if member:
            return member
        self.warn_once(f'member:{member_id}', f'Membre introuvable: {member_id}. Valeur ignoree.')
        return None

    def resolve_event(self, payload):
        if not payload:
            return None
        queryset = Event.objects.filter(
            title=payload['title'],
            start_date=self.to_date(payload['start_date']),
        )
        site_name = payload.get('site_name')
        if site_name:
            queryset = queryset.filter(site__name=site_name)
        event = queryset.first()
        if event:
            return event
        self.warn_once(
            f"event:{payload['title']}:{payload['start_date']}",
            f"Evenement introuvable: {payload['title']} ({payload['start_date']}). Valeur ignoree.",
        )
        return None

    def resolve_finance_category(self, name):
        if not name:
            return None
        category = FinanceCategory.objects.filter(name=name).first()
        if category:
            return category
        raise FinanceImportError(f'Categorie financiere introuvable: {name}')

    def resolve_budget_category(self, name):
        if not name:
            return None
        category = BudgetCategory.objects.filter(name=name).first()
        if category:
            return category
        raise FinanceImportError(f'Categorie de budget introuvable: {name}')

    def resolve_budget_item(self, payload):
        if not payload:
            return None
        group = self.resolve_group(payload.get('group')) if payload.get('group') else None
        department = self.resolve_department(payload.get('department')) if payload.get('department') else None
        category = self.resolve_budget_category(payload['category_name'])

        queryset = BudgetItem.objects.select_related('budget', 'category').filter(
            budget__year=payload['budget_year'],
            category=category,
        )
        if group:
            queryset = queryset.filter(budget__group=group)
        elif department:
            queryset = queryset.filter(budget__department=department)
        else:
            queryset = queryset.filter(budget__name=payload['budget_name'])
        item = queryset.first()
        if item:
            return item
        self.warn_once(
            f"budget_item:{payload['budget_name']}:{payload['category_name']}",
            f"Ligne de budget introuvable pour transaction: {payload['budget_name']} / {payload['category_name']}",
        )
        return None

    def import_finance_categories(self, payloads):
        pending_parents = []
        for payload in payloads:
            obj, created = FinanceCategory.objects.get_or_create(
                name=payload['name'],
                defaults={
                    'description': payload.get('description', ''),
                    'is_income': payload.get('is_income', False),
                    'budget_annual': self.to_decimal(payload.get('budget_annual')),
                    'is_active': payload.get('is_active', True),
                },
            )
            if created:
                self.bump('FinanceCategory', 'created')
            elif self.apply_values(obj, {
                'description': payload.get('description', ''),
                'is_income': payload.get('is_income', False),
                'budget_annual': self.to_decimal(payload.get('budget_annual')),
                'is_active': payload.get('is_active', True),
            }):
                self.bump('FinanceCategory', 'updated')
            pending_parents.append((obj, payload.get('parent_name')))

        for obj, parent_name in pending_parents:
            parent = FinanceCategory.objects.filter(name=parent_name).first() if parent_name else None
            if obj.parent != parent:
                obj.parent = parent
                obj.save()

    def import_budget_categories(self, payloads):
        for payload in payloads:
            obj, created = BudgetCategory.objects.get_or_create(
                name=payload['name'],
                defaults={
                    'description': payload.get('description', ''),
                    'color': payload.get('color', '#0A36FF'),
                    'is_active': payload.get('is_active', True),
                },
            )
            if created:
                self.bump('BudgetCategory', 'created')
            elif self.apply_values(obj, {
                'description': payload.get('description', ''),
                'color': payload.get('color', '#0A36FF'),
                'is_active': payload.get('is_active', True),
            }):
                self.bump('BudgetCategory', 'updated')

    def import_budget_lines(self, payloads):
        for payload in payloads:
            category = self.resolve_finance_category(payload['category_name'])
            obj, created = BudgetLine.objects.get_or_create(
                category=category,
                year=payload['year'],
                month=payload.get('month'),
                defaults={
                    'planned_amount': self.to_decimal(payload['planned_amount']),
                    'notes': payload.get('notes', ''),
                },
            )
            if created:
                self.bump('BudgetLine', 'created')
            elif self.apply_values(obj, {
                'planned_amount': self.to_decimal(payload['planned_amount']),
                'notes': payload.get('notes', ''),
            }):
                self.bump('BudgetLine', 'updated')

    def import_budgets(self, payloads):
        for payload in payloads:
            group = self.resolve_group(payload.get('group')) if payload.get('group') else None
            department = self.resolve_department(payload.get('department')) if payload.get('department') else None
            defaults = {
                'name': payload['name'],
                'total_requested': self.to_decimal(payload.get('total_requested')) or Decimal('0'),
                'total_approved': self.to_decimal(payload.get('total_approved')) or Decimal('0'),
                'status': payload.get('status') or Budget.Status.DRAFT,
                'created_by': self.resolve_user(payload.get('created_by')),
                'approved_by': self.resolve_user(payload.get('approved_by')),
                'submitted_at': self.to_datetime(payload.get('submitted_at')),
                'approved_at': self.to_datetime(payload.get('approved_at')),
                'description': payload.get('description', ''),
                'approval_notes': payload.get('approval_notes', ''),
            }

            if group:
                budget, created = Budget.objects.get_or_create(year=payload['year'], group=group, defaults=defaults)
            elif department:
                budget, created = Budget.objects.get_or_create(year=payload['year'], department=department, defaults=defaults)
            else:
                budget = Budget.objects.filter(
                    year=payload['year'],
                    name=payload['name'],
                    group=None,
                    department=None,
                ).first()
                created = budget is None
                if created:
                    budget = Budget.objects.create(year=payload['year'], group=None, department=None, **defaults)

            if created:
                self.bump('Budget', 'created')
            else:
                values = defaults | {'name': payload['name']}
                if self.apply_values(budget, values):
                    self.bump('Budget', 'updated')

            for item_payload in payload.get('items', []):
                category = self.resolve_budget_category(item_payload['category_name'])
                item_defaults = {
                    'requested_amount': self.to_decimal(item_payload.get('requested_amount')) or Decimal('0'),
                    'approved_amount': self.to_decimal(item_payload.get('approved_amount')) or Decimal('0'),
                    'approval_status': item_payload.get('approval_status', BudgetItem.ApprovalStatus.PENDING),
                    'approved_by': self.resolve_user(item_payload.get('approved_by')),
                    'approved_at': self.to_datetime(item_payload.get('approved_at')),
                    'approval_comments': item_payload.get('approval_comments', ''),
                    'rejection_reason': item_payload.get('rejection_reason', ''),
                    'description': item_payload.get('description', ''),
                    'justification': item_payload.get('justification', ''),
                    'priority': item_payload.get('priority', 1),
                }
                item, item_created = BudgetItem.objects.get_or_create(
                    budget=budget,
                    category=category,
                    defaults=item_defaults,
                )
                if item_created:
                    self.bump('BudgetItem', 'created')
                elif self.apply_values(item, item_defaults):
                    self.bump('BudgetItem', 'updated')

    def import_budget_requests(self, payloads):
        for payload in payloads:
            category = self.resolve_budget_category(payload['category_name'])
            group = self.resolve_group(payload.get('group')) if payload.get('group') else None
            department = self.resolve_department(payload.get('department')) if payload.get('department') else None
            defaults = {
                'description': payload.get('description', ''),
                'approved_amount': self.to_decimal(payload.get('approved_amount')) or Decimal('0'),
                'urgency': payload.get('urgency') or BudgetRequest.Urgency.MEDIUM,
                'needed_by': self.to_date(payload.get('needed_by')),
                'status': payload.get('status') or BudgetRequest.Status.PENDING,
                'requested_by': self.resolve_user(payload.get('requested_by')),
                'reviewed_by': self.resolve_user(payload.get('reviewed_by')),
                'reviewed_at': self.to_datetime(payload.get('reviewed_at')),
                'justification': payload.get('justification', ''),
                'review_notes': payload.get('review_notes', ''),
            }
            obj, created = BudgetRequest.objects.get_or_create(
                title=payload['title'],
                requested_amount=self.to_decimal(payload.get('requested_amount')) or Decimal('0'),
                category=category,
                group=group,
                department=department,
                defaults=defaults,
            )
            if created:
                self.bump('BudgetRequest', 'created')
            elif self.apply_values(obj, defaults | {'title': payload['title']}):
                self.bump('BudgetRequest', 'updated')

    def import_forecasts(self, payloads):
        for payload in payloads:
            forecast, created = BudgetForecast.objects.get_or_create(
                year=payload['year'],
                scenario=payload['scenario'],
                defaults={
                    'name': payload['name'],
                    'description': payload.get('description', ''),
                    'is_active': payload.get('is_active', True),
                    'created_by': self.resolve_user(payload.get('created_by')),
                },
            )
            if created:
                self.bump('BudgetForecast', 'created')
            elif self.apply_values(forecast, {
                'name': payload['name'],
                'description': payload.get('description', ''),
                'is_active': payload.get('is_active', True),
                'created_by': self.resolve_user(payload.get('created_by')),
            }):
                self.bump('BudgetForecast', 'updated')

            forecast.lines.all().delete()
            for line_payload in payload.get('lines', []):
                ForecastLine.objects.create(
                    forecast=forecast,
                    label=line_payload['label'],
                    line_type=line_payload['line_type'],
                    category=self.resolve_finance_category(line_payload.get('category_name')) if line_payload.get('category_name') else None,
                    jan=self.to_decimal(line_payload.get('jan')) or Decimal('0'),
                    feb=self.to_decimal(line_payload.get('feb')) or Decimal('0'),
                    mar=self.to_decimal(line_payload.get('mar')) or Decimal('0'),
                    apr=self.to_decimal(line_payload.get('apr')) or Decimal('0'),
                    may=self.to_decimal(line_payload.get('may')) or Decimal('0'),
                    jun=self.to_decimal(line_payload.get('jun')) or Decimal('0'),
                    jul=self.to_decimal(line_payload.get('jul')) or Decimal('0'),
                    aug=self.to_decimal(line_payload.get('aug')) or Decimal('0'),
                    sep=self.to_decimal(line_payload.get('sep')) or Decimal('0'),
                    oct=self.to_decimal(line_payload.get('oct')) or Decimal('0'),
                    nov=self.to_decimal(line_payload.get('nov')) or Decimal('0'),
                    dec=self.to_decimal(line_payload.get('dec')) or Decimal('0'),
                    notes=line_payload.get('notes', ''),
                )
                self.bump('ForecastLine', 'created')

    def import_transactions(self, payloads):
        for payload in payloads:
            defaults = {
                'site': self.resolve_site(payload.get('site_name')),
                'amount': self.to_decimal(payload['amount']),
                'transaction_type': payload['transaction_type'],
                'payment_method': payload.get('payment_method') or FinancialTransaction.PaymentMethod.ESPECES,
                'status': payload.get('status') or FinancialTransaction.Status.EN_ATTENTE,
                'transaction_date': self.to_date(payload['transaction_date']),
                'description': payload.get('description', ''),
                'category': self.resolve_finance_category(payload.get('category_name')) if payload.get('category_name') else None,
                'member': self.resolve_member(payload.get('member_id')),
                'event': self.resolve_event(payload.get('event')),
                'budget_item': self.resolve_budget_item(payload.get('budget_item')),
                'recorded_by': self.resolve_user(payload.get('recorded_by')),
                'validated_by': self.resolve_user(payload.get('validated_by')),
                'validated_at': self.to_datetime(payload.get('validated_at')),
                'notes': payload.get('notes', ''),
                'is_deleted': payload.get('is_deleted', False),
                'deleted_at': self.to_datetime(payload.get('deleted_at')),
                'deleted_by': self.resolve_user(payload.get('deleted_by')),
            }
            obj, created = FinancialTransaction.all_objects.get_or_create(
                reference=payload['reference'],
                defaults=defaults,
            )
            if created:
                self.bump('FinancialTransaction', 'created')
            elif self.apply_values(obj, defaults):
                self.bump('FinancialTransaction', 'updated')