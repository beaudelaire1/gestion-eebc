"""
Vues pour le système de budget.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import date, datetime
from decimal import Decimal

from .models import Budget, BudgetItem, BudgetCategory, BudgetRequest, FinancialTransaction
from apps.groups.models import Group
from apps.departments.models import Department


@login_required
def budget_dashboard(request):
    """Tableau de bord des budgets."""
    current_year = date.today().year
    
    # Statistiques générales
    active_budgets = Budget.objects.filter(
        year=current_year,
        status=Budget.Status.ACTIVE
    )
    
    total_approved = active_budgets.aggregate(
        total=Sum('total_approved')
    )['total'] or Decimal('0.00')
    
    total_spent = sum(budget.spent_amount for budget in active_budgets)
    
    pending_requests = BudgetRequest.objects.filter(
        status=BudgetRequest.Status.PENDING
    ).count()
    
    # Budgets par statut
    budget_stats = {}
    for status_code, status_name in Budget.Status.choices:
        count = Budget.objects.filter(
            year=current_year,
            status=status_code
        ).count()
        budget_stats[status_name] = count
    
    # Demandes urgentes
    urgent_requests = BudgetRequest.objects.filter(
        status=BudgetRequest.Status.PENDING,
        urgency__in=[BudgetRequest.Urgency.HIGH, BudgetRequest.Urgency.URGENT]
    ).order_by('-created_at')[:5]
    
    # Budgets avec faible reste
    low_budget_items = []
    for budget in active_budgets:
        for item in budget.items.all():
            if item.approved_amount > 0:
                utilization = item.utilization_percentage
                if utilization > 80:  # Plus de 80% utilisé
                    low_budget_items.append(item)
    
    context = {
        'current_year': current_year,
        'total_approved': total_approved,
        'total_spent': total_spent,
        'remaining_budget': total_approved - total_spent,
        'utilization_percentage': (total_spent / total_approved * 100) if total_approved > 0 else 0,
        'pending_requests': pending_requests,
        'budget_stats': budget_stats,
        'urgent_requests': urgent_requests,
        'low_budget_items': low_budget_items[:5],
    }
    
    return render(request, 'finance/budget/dashboard.html', context)


@login_required
def budget_list(request):
    """Liste des budgets."""
    year = request.GET.get('year', date.today().year)
    status = request.GET.get('status', '')
    
    budgets = Budget.objects.filter(year=year)
    
    if status:
        budgets = budgets.filter(status=status)
    
    # Calculer les statistiques pour chaque budget
    budget_data = []
    for budget in budgets:
        budget_data.append({
            'budget': budget,
            'spent_amount': budget.spent_amount,
            'remaining_amount': budget.remaining_amount,
            'utilization_percentage': budget.utilization_percentage,
        })
    
    # Années disponibles
    available_years = Budget.objects.values_list('year', flat=True).distinct().order_by('-year')
    
    context = {
        'budget_data': budget_data,
        'current_year': int(year),
        'current_status': status,
        'available_years': available_years,
        'status_choices': Budget.Status.choices,
    }
    
    return render(request, 'finance/budget/list.html', context)


@login_required
def budget_detail(request, budget_id):
    """Détail d'un budget."""
    budget = get_object_or_404(Budget, id=budget_id)
    
    # Vérifier les permissions
    user_can_edit = (
        request.user.is_admin or
        request.user.role in ['admin', 'finance'] or
        budget.created_by == request.user
    )
    
    # Statistiques par catégorie
    items_data = []
    for item in budget.items.all():
        items_data.append({
            'item': item,
            'spent_amount': item.spent_amount,
            'remaining_amount': item.remaining_amount,
            'utilization_percentage': item.utilization_percentage,
        })
    
    # Transactions liées
    from apps.finance.models import FinancialTransaction
    transactions = FinancialTransaction.objects.filter(
        budget_item__budget=budget
    ).order_by('-transaction_date')[:10]
    
    context = {
        'budget': budget,
        'items_data': items_data,
        'transactions': transactions,
        'user_can_edit': user_can_edit,
    }
    
    return render(request, 'finance/budget/detail.html', context)


@login_required
def budget_create(request):
    """Créer un nouveau budget."""
    if not (request.user.is_admin or request.user.role in ['admin', 'finance']):
        messages.error(request, 'Vous n\'avez pas les permissions pour créer un budget.')
        return redirect('finance:budget_list')
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        name = request.POST.get('name')
        year = int(request.POST.get('year'))
        description = request.POST.get('description', '')
        
        # Entité (groupe ou département)
        group_id = request.POST.get('group')
        department_id = request.POST.get('department')
        
        if not (group_id or department_id):
            messages.error(request, 'Vous devez sélectionner un groupe ou un département.')
            return render(request, 'finance/budget/create.html', {
                'groups': Group.objects.filter(is_active=True),
                'departments': Department.objects.filter(is_active=True),
                'categories': BudgetCategory.objects.filter(is_active=True),
            })
        
        # Créer le budget
        budget = Budget.objects.create(
            name=name,
            year=year,
            description=description,
            group_id=group_id if group_id else None,
            department_id=department_id if department_id else None,
            created_by=request.user,
            total_requested=Decimal('0.00')
        )
        
        # Créer les lignes de budget à partir du tableau libre
        total_requested = Decimal('0.00')
        
        # Traiter les lignes du tableau libre
        line_names = {}
        line_descriptions = {}
        line_amounts = {}
        line_priorities = {}
        
        # Regrouper les données par ligne
        for key, value in request.POST.items():
            if key.startswith('line_'):
                parts = key.split('_')
                if len(parts) >= 3:
                    field_type = parts[1]  # name, description, amount, priority
                    line_id = parts[2]
                    
                    if field_type == 'name':
                        line_names[line_id] = value
                    elif field_type == 'description':
                        line_descriptions[line_id] = value
                    elif field_type == 'amount' and value:
                        line_amounts[line_id] = Decimal(value)
                    elif field_type == 'priority':
                        line_priorities[line_id] = int(value)
        
        # Créer les lignes de budget
        for line_id in line_names.keys():
            if line_id in line_amounts and line_amounts[line_id] > 0:
                # Créer une catégorie personnalisée pour cette ligne
                category_name = f"Libre - {line_names[line_id]}"
                custom_category, created = BudgetCategory.objects.get_or_create(
                    name=category_name,
                    defaults={
                        'description': f"Ligne libre: {line_names[line_id]}",
                        'color': '#28A745',  # Vert pour les lignes libres
                        'is_active': True
                    }
                )
                
                BudgetItem.objects.create(
                    budget=budget,
                    category=custom_category,
                    requested_amount=line_amounts[line_id],
                    description=line_descriptions.get(line_id, ''),
                    justification=f"Ligne libre: {line_names[line_id]}",
                    priority=line_priorities.get(line_id, 3)
                )
                
                total_requested += line_amounts[line_id]
        
        # Mettre à jour le total
        budget.total_requested = total_requested
        budget.save()
        
        messages.success(request, f'Budget "{budget.name}" créé avec succès.')
        return redirect('finance:budget_detail', budget_id=budget.id)
    
    context = {
        'groups': Group.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        'categories': BudgetCategory.objects.filter(is_active=True),
        'current_year': date.today().year,
    }
    
    return render(request, 'finance/budget/create_simple.html', context)


@login_required
def budget_request_list(request):
    """Liste des demandes de budget."""
    status = request.GET.get('status', '')
    urgency = request.GET.get('urgency', '')
    
    requests = BudgetRequest.objects.all()
    
    if status:
        requests = requests.filter(status=status)
    
    if urgency:
        requests = requests.filter(urgency=urgency)
    
    requests = requests.order_by('-created_at')
    
    context = {
        'requests': requests,
        'current_status': status,
        'current_urgency': urgency,
        'status_choices': BudgetRequest.Status.choices,
        'urgency_choices': BudgetRequest.Urgency.choices,
    }
    
    return render(request, 'finance/budget/request_list.html', context)


@login_required
def budget_request_create(request):
    """Créer une demande de budget."""
    if request.method == 'POST':
        # Récupérer les données
        title = request.POST.get('title')
        description = request.POST.get('description')
        justification = request.POST.get('justification')
        requested_amount = Decimal(request.POST.get('requested_amount'))
        category_id = request.POST.get('category')
        urgency = request.POST.get('urgency')
        needed_by = request.POST.get('needed_by')
        
        # Entité
        group_id = request.POST.get('group')
        department_id = request.POST.get('department')
        
        # Créer la demande
        budget_request = BudgetRequest.objects.create(
            title=title,
            description=description,
            justification=justification,
            requested_amount=requested_amount,
            category_id=category_id,
            urgency=urgency,
            needed_by=datetime.strptime(needed_by, '%Y-%m-%d').date() if needed_by else None,
            group_id=group_id if group_id else None,
            department_id=department_id if department_id else None,
            requested_by=request.user
        )
        
        messages.success(request, 'Demande de budget créée avec succès.')
        return redirect('finance:budget_request_detail', request_id=budget_request.id)
    
    context = {
        'groups': Group.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        'categories': BudgetCategory.objects.filter(is_active=True),
        'urgency_choices': BudgetRequest.Urgency.choices,
    }
    
    return render(request, 'finance/budget/request_create.html', context)


@login_required
def budget_request_detail(request, request_id):
    """Détail d'une demande de budget."""
    budget_request = get_object_or_404(BudgetRequest, id=request_id)
    
    # Vérifier les permissions
    user_can_review = (
        request.user.is_admin or
        request.user.role in ['admin', 'finance']
    )
    
    context = {
        'budget_request': budget_request,
        'user_can_review': user_can_review,
    }
    
    return render(request, 'finance/budget/request_detail.html', context)


@login_required
def budget_approve_detailed(request, budget_id):
    """Approbation détaillée ligne par ligne."""
    if not (request.user.is_admin or request.user.role in ['admin', 'finance']):
        messages.error(request, 'Vous n\'avez pas les permissions pour approuver un budget.')
        return redirect('finance:budget_detail', budget_id=budget_id)
    
    budget = get_object_or_404(Budget, id=budget_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve_item':
            # Approuver une ligne spécifique
            item_id = request.POST.get('item_id')
            item = get_object_or_404(BudgetItem, id=item_id, budget=budget)
            
            approved_amount = Decimal(request.POST.get('approved_amount', '0'))
            comments = request.POST.get('comments', '')
            
            item.approved_amount = approved_amount
            item.approval_status = BudgetItem.ApprovalStatus.APPROVED if approved_amount > 0 else BudgetItem.ApprovalStatus.REJECTED
            item.approved_by = request.user
            item.approved_at = timezone.now()
            item.approval_comments = comments
            item.save()
            
            messages.success(request, f'Ligne "{item.category.name}" approuvée.')
            
        elif action == 'reject_item':
            # Rejeter une ligne spécifique
            item_id = request.POST.get('item_id')
            item = get_object_or_404(BudgetItem, id=item_id, budget=budget)
            
            rejection_reason = request.POST.get('rejection_reason', '')
            
            item.approved_amount = Decimal('0')
            item.approval_status = BudgetItem.ApprovalStatus.REJECTED
            item.approved_by = request.user
            item.approved_at = timezone.now()
            item.rejection_reason = rejection_reason
            item.save()
            
            messages.warning(request, f'Ligne "{item.category.name}" rejetée.')
            
        elif action == 'finalize_budget':
            # Finaliser l'approbation du budget
            budget_comments = request.POST.get('budget_comments', '')
            
            # Calculer le total approuvé
            total_approved = sum(item.approved_amount for item in budget.items.all())
            
            budget.total_approved = total_approved
            budget.status = Budget.Status.APPROVED
            budget.approved_by = request.user
            budget.approved_at = timezone.now()
            budget.approval_notes = budget_comments
            budget.save()
            
            messages.success(request, f'Budget "{budget.name}" finalisé et approuvé.')
            return redirect('finance:budget_detail', budget_id=budget.id)
        
        return redirect('finance:budget_approve_detailed', budget_id=budget.id)
    
    # Statistiques d'approbation
    items_stats = {
        'total': budget.items.count(),
        'pending': budget.items.filter(approval_status=BudgetItem.ApprovalStatus.PENDING).count(),
        'approved': budget.items.filter(approval_status=BudgetItem.ApprovalStatus.APPROVED).count(),
        'rejected': budget.items.filter(approval_status=BudgetItem.ApprovalStatus.REJECTED).count(),
    }
    
    context = {
        'budget': budget,
        'items_stats': items_stats,
    }
    
    return render(request, 'finance/budget/approve_detailed.html', context)

@login_required
def budget_submit(request, budget_id):
    """Soumettre un budget pour approbation."""
    budget = get_object_or_404(Budget, id=budget_id)
    
    # Vérifier les permissions
    if not (request.user == budget.created_by or request.user.is_admin or request.user.role in ['admin', 'finance']):
        messages.error(request, 'Vous n\'avez pas les permissions pour soumettre ce budget.')
        return redirect('finance:budget_detail', budget_id=budget_id)
    
    if budget.status != Budget.Status.DRAFT:
        messages.error(request, 'Seuls les budgets en brouillon peuvent être soumis.')
        return redirect('finance:budget_detail', budget_id=budget_id)
    
    if request.method == 'POST':
        # Vérifier qu'il y a au moins une ligne
        if not budget.items.exists():
            messages.error(request, 'Le budget doit contenir au moins une ligne.')
            return redirect('finance:budget_detail', budget_id=budget_id)
        
        budget.status = Budget.Status.SUBMITTED
        budget.submitted_at = timezone.now()
        budget.save()
        
        messages.success(request, f'Budget "{budget.name}" soumis pour approbation.')
        return redirect('finance:budget_detail', budget_id=budget.id)
    
    return render(request, 'finance/budget/submit_confirm.html', {'budget': budget})


@login_required
def budget_edit(request, budget_id):
    """Éditer un budget en brouillon."""
    budget = get_object_or_404(Budget, id=budget_id)
    
    # Vérifier les permissions
    if not (request.user == budget.created_by or request.user.is_admin or request.user.role in ['admin', 'finance']):
        messages.error(request, 'Vous n\'avez pas les permissions pour éditer ce budget.')
        return redirect('finance:budget_detail', budget_id=budget_id)
    
    if budget.status != Budget.Status.DRAFT:
        messages.error(request, 'Seuls les budgets en brouillon peuvent être édités.')
        return redirect('finance:budget_detail', budget_id=budget_id)
    
    if request.method == 'POST':
        # Traitement en mode édition - CONSERVER les lignes existantes
        # Ne PAS supprimer les anciennes lignes, les mettre à jour
        
        # Traiter les lignes du tableau libre
        total_requested = Decimal('0.00')
        
        line_names = {}
        line_descriptions = {}
        line_amounts = {}
        line_priorities = {}
        line_ids = {}  # Pour mapper les IDs existants
        
        # Regrouper les données par ligne
        for key, value in request.POST.items():
            if key.startswith('line_'):
                parts = key.split('_')
                if len(parts) >= 3:
                    field_type = parts[1]
                    line_id = parts[2]
                    
                    if field_type == 'name':
                        line_names[line_id] = value
                    elif field_type == 'description':
                        line_descriptions[line_id] = value
                    elif field_type == 'amount' and value:
                        line_amounts[line_id] = Decimal(value)
                    elif field_type == 'priority':
                        line_priorities[line_id] = int(value)
                    elif field_type == 'id' and value:  # ID de ligne existante
                        line_ids[line_id] = int(value)
        
        # Supprimer seulement les lignes qui ne sont plus dans le formulaire
        existing_item_ids = set(budget.items.values_list('id', flat=True))
        submitted_item_ids = set(line_ids.values())
        items_to_delete = existing_item_ids - submitted_item_ids
        
        if items_to_delete:
            BudgetItem.objects.filter(id__in=items_to_delete).delete()
        
        # Traiter chaque ligne (mise à jour ou création)
        for line_id in line_names.keys():
            if line_id in line_amounts and line_amounts[line_id] > 0:
                category_name = f"Libre - {line_names[line_id]}"
                custom_category, created = BudgetCategory.objects.get_or_create(
                    name=category_name,
                    defaults={
                        'description': f"Ligne libre: {line_names[line_id]}",
                        'color': '#28A745',
                        'is_active': True
                    }
                )
                
                # Vérifier si c'est une ligne existante ou nouvelle
                if line_id in line_ids:
                    # Mettre à jour la ligne existante
                    try:
                        item = BudgetItem.objects.get(id=line_ids[line_id], budget=budget)
                        item.category = custom_category
                        item.requested_amount = line_amounts[line_id]
                        item.description = line_descriptions.get(line_id, '')
                        item.justification = f"Ligne libre: {line_names[line_id]}"
                        item.priority = line_priorities.get(line_id, 3)
                        item.save()
                    except BudgetItem.DoesNotExist:
                        # Créer une nouvelle ligne si l'ID n'existe pas
                        BudgetItem.objects.create(
                            budget=budget,
                            category=custom_category,
                            requested_amount=line_amounts[line_id],
                            description=line_descriptions.get(line_id, ''),
                            justification=f"Ligne libre: {line_names[line_id]}",
                            priority=line_priorities.get(line_id, 3)
                        )
                else:
                    # Créer une nouvelle ligne
                    BudgetItem.objects.create(
                        budget=budget,
                        category=custom_category,
                        requested_amount=line_amounts[line_id],
                        description=line_descriptions.get(line_id, ''),
                        justification=f"Ligne libre: {line_names[line_id]}",
                        priority=line_priorities.get(line_id, 3)
                    )
                
                total_requested += line_amounts[line_id]
        
        # Mettre à jour le budget
        budget.name = request.POST.get('name', budget.name)
        budget.description = request.POST.get('description', budget.description)
        budget.total_requested = total_requested
        budget.save()
        
        messages.success(request, f'Budget "{budget.name}" mis à jour avec succès.')
        return redirect('finance:budget_detail', budget_id=budget.id)
    
    # Préparer les données pour l'édition
    context = {
        'budget': budget,
        'groups': Group.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        'current_year': date.today().year,
        'edit_mode': True,
    }
    
    return render(request, 'finance/budget/edit.html', context)

# =============================================================================
# FONCTIONNALITÉS D'EXPORT ET D'IMPRESSION
# =============================================================================

from django.http import HttpResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime


@login_required
def budget_export_excel(request, budget_id):
    """Exporter un budget en Excel."""
    budget = get_object_or_404(Budget, id=budget_id)
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Budget {budget.year}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-tête du document
    ws.merge_cells('A1:G1')
    ws['A1'] = f"BUDGET {budget.year} - {budget.entity}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informations générales
    row = 3
    ws[f'A{row}'] = "Nom du budget:"
    ws[f'B{row}'] = budget.name
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Année:"
    ws[f'B{row}'] = budget.year
    
    row += 1
    ws[f'A{row}'] = "Entité:"
    ws[f'B{row}'] = str(budget.entity)
    
    row += 1
    ws[f'A{row}'] = "Statut:"
    ws[f'B{row}'] = budget.get_status_display()
    
    row += 1
    ws[f'A{row}'] = "Créé par:"
    ws[f'B{row}'] = str(budget.created_by)
    
    row += 1
    ws[f'A{row}'] = "Date de création:"
    ws[f'B{row}'] = budget.created_at.strftime('%d/%m/%Y')
    
    # Résumé financier
    row += 2
    ws[f'A{row}'] = "RÉSUMÉ FINANCIER"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 1
    ws[f'A{row}'] = "Montant demandé:"
    ws[f'B{row}'] = f"{budget.total_requested} €"
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Montant approuvé:"
    ws[f'B{row}'] = f"{budget.total_approved} €"
    ws[f'B{row}'].font = Font(bold=True, color="008000")
    
    row += 1
    ws[f'A{row}'] = "Montant dépensé:"
    ws[f'B{row}'] = f"{budget.spent_amount} €"
    
    row += 1
    ws[f'A{row}'] = "Montant restant:"
    ws[f'B{row}'] = f"{budget.remaining_amount} €"
    
    # En-têtes du tableau des lignes
    row += 3
    headers = ['Catégorie', 'Description', 'Priorité', 'Montant Demandé', 'Montant Approuvé', 'Montant Dépensé', 'Restant']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Données des lignes de budget
    for item in budget.items.all():
        row += 1
        data = [
            item.category.name,
            item.description,
            f"Priorité {item.priority}",
            f"{item.requested_amount} €",
            f"{item.approved_amount} €",
            f"{item.spent_amount} €",
            f"{item.remaining_amount} €"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col >= 4:  # Colonnes de montants
                cell.alignment = Alignment(horizontal='right')
    
    # Ligne de total
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{budget.total_requested} €").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{budget.total_approved} €").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"{budget.spent_amount} €").font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"{budget.remaining_amount} €").font = Font(bold=True)
    
    # Ajuster la largeur des colonnes
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="budget_{budget.year}_{budget.entity}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    # Sauvegarder le fichier Excel dans la réponse
    wb.save(response)
    return response


@login_required
def budget_print_view(request, budget_id):
    """Vue d'impression pour un budget."""
    budget = get_object_or_404(Budget, id=budget_id)
    
    # Calculer les statistiques pour l'impression
    items_data = []
    for item in budget.items.all():
        items_data.append({
            'item': item,
            'spent_amount': item.spent_amount,
            'remaining_amount': item.remaining_amount,
            'utilization_percentage': item.utilization_percentage,
        })
    
    context = {
        'budget': budget,
        'items_data': items_data,
        'print_date': datetime.now(),
        'total_spent': budget.spent_amount,
        'total_remaining': budget.remaining_amount,
        'utilization_percentage': budget.utilization_percentage,
    }
    
    return render(request, 'finance/budget/print.html', context)


@login_required
def budget_list_export_excel(request):
    """Exporter la liste des budgets en Excel."""
    year = request.GET.get('year', date.today().year)
    status = request.GET.get('status', '')
    
    budgets = Budget.objects.filter(year=year)
    if status:
        budgets = budgets.filter(status=status)
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Budgets {year}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells('A1:I1')
    ws['A1'] = f"LISTE DES BUDGETS {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # En-têtes
    row = 3
    headers = ['Nom', 'Entité', 'Statut', 'Montant Demandé', 'Montant Approuvé', 'Montant Dépensé', 'Restant', 'Utilisation %', 'Créé le']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for budget in budgets:
        row += 1
        data = [
            budget.name,
            str(budget.entity),
            budget.get_status_display(),
            f"{budget.total_requested} €",
            f"{budget.total_approved} €",
            f"{budget.spent_amount} €",
            f"{budget.remaining_amount} €",
            f"{budget.utilization_percentage:.1f}%",
            budget.created_at.strftime('%d/%m/%Y')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [4, 5, 6, 7]:  # Colonnes de montants
                cell.alignment = Alignment(horizontal='right')
    
    # Ajuster la largeur des colonnes
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="budgets_{year}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def transactions_export_excel(request):
    """Exporter les transactions financières en Excel."""
    # Filtres
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('type')
    status = request.GET.get('status')
    
    transactions = FinancialTransaction.objects.all()
    
    if start_date:
        transactions = transactions.filter(transaction_date__gte=start_date)
    if end_date:
        transactions = transactions.filter(transaction_date__lte=end_date)
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if status:
        transactions = transactions.filter(status=status)
    
    transactions = transactions.order_by('-transaction_date')
    
    # Créer le classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre
    ws.merge_cells('A1:J1')
    ws['A1'] = "TRANSACTIONS FINANCIÈRES"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Filtres appliqués
    row = 2
    if start_date or end_date:
        period = f"Période: {start_date or 'début'} - {end_date or 'fin'}"
        ws[f'A{row}'] = period
        row += 1
    
    # En-têtes
    row += 1
    headers = ['Référence', 'Date', 'Type', 'Montant', 'Méthode', 'Statut', 'Description', 'Catégorie', 'Membre', 'Enregistré par']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    total_income = Decimal('0.00')
    total_expense = Decimal('0.00')
    
    for transaction in transactions:
        row += 1
        data = [
            transaction.reference,
            transaction.transaction_date.strftime('%d/%m/%Y'),
            transaction.get_transaction_type_display(),
            f"{transaction.amount} €",
            transaction.get_payment_method_display(),
            transaction.get_status_display(),
            transaction.description,
            str(transaction.category) if transaction.category else '',
            str(transaction.member) if transaction.member else '',
            str(transaction.recorded_by) if transaction.recorded_by else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 4:  # Colonne montant
                cell.alignment = Alignment(horizontal='right')
        
        # Calculer les totaux
        if transaction.is_income:
            total_income += transaction.amount
        else:
            total_expense += transaction.amount
    
    # Totaux
    row += 2
    ws[f'A{row}'] = "TOTAUX"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total Entrées:"
    ws[f'B{row}'] = f"{total_income} €"
    ws[f'B{row}'].font = Font(bold=True, color="008000")
    
    row += 1
    ws[f'A{row}'] = "Total Sorties:"
    ws[f'B{row}'] = f"{total_expense} €"
    ws[f'B{row}'].font = Font(bold=True, color="FF0000")
    
    row += 1
    ws[f'A{row}'] = "Solde:"
    ws[f'B{row}'] = f"{total_income - total_expense} €"
    ws[f'B{row}'].font = Font(bold=True)
    
    # Ajuster la largeur des colonnes
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response